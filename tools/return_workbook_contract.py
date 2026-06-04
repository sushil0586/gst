#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SheetContract:
    title: str
    dimensions: str
    first_non_empty_row_index: int | None
    first_non_empty_row: list[str]
    sample_rows: list[list[str]]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def inspect_workbook(path: Path, max_rows: int = 6) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[SheetContract] = []

    for worksheet in workbook.worksheets:
        sample_rows: list[list[str]] = []
        first_non_empty_row_index = None
        first_non_empty_row: list[str] = []

        for index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True),
            start=1,
        ):
            normalized = [normalize_cell(value) for value in row]
            sample_rows.append(normalized)
            if first_non_empty_row_index is None and any(cell != "" for cell in normalized):
                first_non_empty_row_index = index
                first_non_empty_row = normalized

        sheets.append(
            SheetContract(
                title=worksheet.title,
                dimensions=worksheet.calculate_dimension(),
                first_non_empty_row_index=first_non_empty_row_index,
                first_non_empty_row=first_non_empty_row,
                sample_rows=sample_rows,
            )
        )

    return {
        "file": str(path),
        "sheet_names": [sheet.title for sheet in sheets],
        "sheet_count": len(sheets),
        "sheets": [asdict(sheet) for sheet in sheets],
    }


def compare_workbooks(base: dict[str, Any], target: dict[str, Any]) -> dict[str, Any]:
    base_sheets = {sheet["title"]: sheet for sheet in base["sheets"]}
    target_sheets = {sheet["title"]: sheet for sheet in target["sheets"]}

    base_names = set(base_sheets)
    target_names = set(target_sheets)

    common = sorted(base_names & target_names)
    only_in_base = sorted(base_names - target_names)
    only_in_target = sorted(target_names - base_names)

    header_differences = []
    for name in common:
        base_header = base_sheets[name]["first_non_empty_row"]
        target_header = target_sheets[name]["first_non_empty_row"]
        if base_header != target_header:
            header_differences.append(
                {
                    "sheet": name,
                    "base_header": base_header,
                    "target_header": target_header,
                }
            )

    return {
        "base_file": base["file"],
        "target_file": target["file"],
        "common_sheets": common,
        "only_in_base": only_in_base,
        "only_in_target": only_in_target,
        "header_differences": header_differences,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect and compare GST return workbook contracts.")
    parser.add_argument("workbook", type=Path, help="Workbook path to inspect.")
    parser.add_argument("--compare", type=Path, help="Optional second workbook to compare against.")
    parser.add_argument("--rows", type=int, default=6, help="Number of sample rows to inspect per sheet.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of human-readable output.")
    args = parser.parse_args()

    base = inspect_workbook(args.workbook, max_rows=args.rows)
    if args.compare:
        target = inspect_workbook(args.compare, max_rows=args.rows)
        payload: dict[str, Any] = {
            "base": base,
            "target": target,
            "comparison": compare_workbooks(base, target),
        }
    else:
        payload = base

    if args.json:
        print(json.dumps(payload, indent=2))
        return

    print(f"Workbook: {base['file']}")
    print(f"Sheets ({base['sheet_count']}): {', '.join(base['sheet_names'])}")
    for sheet in base["sheets"]:
        print(f"- {sheet['title']} [{sheet['dimensions']}]")
        print(f"  first non-empty row ({sheet['first_non_empty_row_index']}): {sheet['first_non_empty_row']}")

    if args.compare:
        comparison = payload["comparison"]
        print("\nComparison")
        print(f"Base only: {comparison['only_in_base']}")
        print(f"Target only: {comparison['only_in_target']}")
        print(f"Common: {comparison['common_sheets']}")
        if comparison["header_differences"]:
            print("Header differences:")
            for diff in comparison["header_differences"]:
                print(f"  - {diff['sheet']}")
                print(f"    base:   {diff['base_header']}")
                print(f"    target: {diff['target_header']}")


if __name__ == "__main__":
    main()
