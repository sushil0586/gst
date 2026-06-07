import json
from decimal import Decimal
from tempfile import SpooledTemporaryFile
from typing import Iterable, Sequence
from uuid import UUID

from django.db.models import QuerySet
from django.http import FileResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.audit_logs.models import AuditLog
from apps.common.services.dashboard import build_close_manager_dashboard, build_close_manager_report
from apps.filings.models import ReturnFiling
from apps.filings.serializers import ReturnFilingSerializer
from apps.gst_transactions.models import GSTTransaction
from apps.imports.models import ImportRowError
from apps.reconciliation.models import ReconciliationItem
from apps.returns.models import ReturnPreparation
from apps.common.services.return_workbooks import export_gstr1_workbook, export_gstr3b_workbook, export_gstr7_workbook, export_gstr9_workbook

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_xlsx_response(*, sheet_title: str, headers: list[str], rows: Iterable[Sequence[object]], filename: str) -> HttpResponse:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_title[:31])
    append_sheet(worksheet=worksheet, title=sheet_title, headers=headers, rows=rows)
    return finalize_workbook_response(workbook=workbook, filename=filename)


def build_multi_sheet_xlsx_response(*, sheets: list[dict[str, object]], filename: str) -> HttpResponse:
    workbook = Workbook(write_only=True)

    for sheet in sheets:
        worksheet = workbook.create_sheet(title=str(sheet["title"])[:31])
        append_sheet(
            worksheet=worksheet,
            title=str(sheet["title"]),
            headers=list(sheet["headers"]),
            rows=sheet["rows"],
        )
    return finalize_workbook_response(workbook=workbook, filename=filename)


def export_transactions(queryset: QuerySet[GSTTransaction]) -> HttpResponse:
    rows = (
        [
            transaction.transaction_type,
            transaction.document_type,
            transaction.reference_number,
            transaction.transaction_date.isoformat() if transaction.transaction_date else "",
            transaction.counterparty_gstin,
            transaction.counterparty_name,
            format_decimal(transaction.taxable_value),
            format_decimal(transaction.cgst_amount),
            format_decimal(transaction.sgst_amount),
            format_decimal(transaction.igst_amount),
            format_decimal(transaction.cess_amount),
            format_decimal(transaction.total_amount),
            transaction.import_batch.file_name if transaction.import_batch else "",
            transaction.status,
        ]
        for transaction in queryset.iterator(chunk_size=2000)
    )
    return build_xlsx_response(
        sheet_title="Transactions",
        headers=[
            "Transaction Type",
            "Document Type",
            "Document Number",
            "Document Date",
            "Counterparty GSTIN",
            "Counterparty Name",
            "Taxable Value",
            "CGST",
            "SGST",
            "IGST",
            "Cess",
            "Total Amount",
            "Source Import Batch",
            "Status",
        ],
        rows=rows,
        filename="gst-transactions.xlsx",
    )


def export_import_errors(queryset: QuerySet[ImportRowError]) -> HttpResponse:
    rows = (
        [
            error.import_batch.file_name,
            error.row_number,
            error.field_name,
            error.error_code,
            error.error_message,
            json.dumps(error.raw_row, ensure_ascii=True),
        ]
        for error in queryset.iterator(chunk_size=2000)
    )
    return build_xlsx_response(
        sheet_title="Import Errors",
        headers=[
            "Import Batch",
            "Row Number",
            "Field",
            "Error Code",
            "Error Message",
            "Raw Data",
        ],
        rows=rows,
        filename="import-errors.xlsx",
    )


def export_reconciliation(queryset: QuerySet[ReconciliationItem]) -> HttpResponse:
    def iter_rows():
        for item in queryset.iterator(chunk_size=2000):
            books_transaction = item.books_transaction
            portal_transaction = item.portal_transaction
            yield [
                item.reconciliation_run_id,
                books_transaction.counterparty_name if books_transaction else (portal_transaction.counterparty_name if portal_transaction else ""),
                books_transaction.counterparty_gstin if books_transaction else (portal_transaction.counterparty_gstin if portal_transaction else ""),
                books_transaction.reference_number if books_transaction else "",
                books_transaction.transaction_date.isoformat() if books_transaction and books_transaction.transaction_date else "",
                format_decimal(books_transaction.tax_amount) if books_transaction else "",
                portal_transaction.reference_number if portal_transaction else "",
                portal_transaction.transaction_date.isoformat() if portal_transaction and portal_transaction.transaction_date else "",
                format_decimal(portal_transaction.tax_amount) if portal_transaction else "",
                item.match_status,
                item.mismatch_reason,
                format_decimal(item.tax_difference),
                item.action_status,
                item.remarks,
            ]

    return build_xlsx_response(
        sheet_title="Reconciliation",
        headers=[
            "Run",
            "Vendor",
            "GSTIN",
            "Books Invoice",
            "Books Date",
            "Books Tax",
            "2B Invoice",
            "2B Date",
            "2B Tax",
            "Match Status",
            "Reason",
            "Difference",
            "Action Status",
            "Remarks",
        ],
        rows=iter_rows(),
        filename="reconciliation-report.xlsx",
    )


def export_return_summary(queryset: QuerySet[ReturnPreparation]) -> HttpResponse:
    def iter_rows():
        for prepared_return in queryset.iterator(chunk_size=1000):
            summary = prepared_return.summary_snapshot or {}
            outward_supplies = summary.get("outward_supplies", {})
            itc_summary = summary.get("itc_summary", {})
            yield [
                prepared_return.return_type,
                prepared_return.status,
                format_decimal(
                    outward_supplies.get("total_taxable_value")
                    or outward_supplies.get("outward_taxable_value")
                    or 0
                ),
                format_decimal(
                    outward_supplies.get("total_tax_amount")
                    or outward_supplies.get("outward_tax_liability")
                    or 0
                ),
                format_decimal(itc_summary.get("eligible_itc") or 0),
                format_decimal(itc_summary.get("net_tax_payable") or 0),
                json.dumps(summary, ensure_ascii=True),
                display_user(prepared_return.prepared_by),
                display_user(prepared_return.approved_by),
                display_user(prepared_return.filed_by),
                prepared_return.filed_at.isoformat() if prepared_return.filed_at else "",
                prepared_return.arn,
            ]

    return build_xlsx_response(
        sheet_title="Return Summary",
        headers=[
            "Return Type",
            "Status",
            "Taxable Value",
            "Tax Amount",
            "Eligible ITC",
            "Net Payable",
            "Summary Fields",
            "Prepared By",
            "Approved By",
            "Filed By",
            "Filed At",
            "ARN",
        ],
        rows=iter_rows(),
        filename="return-summary.xlsx",
    )


def export_filing_evidence_pack(*, filing: ReturnFiling) -> HttpResponse:
    serialized = ReturnFilingSerializer(instance=filing).data
    latest_attempt = serialized.get("latest_attempt") or {}
    support_status = serialized.get("support_status_summary") or {}
    support_actions = serialized.get("support_actions_summary") or {}
    provider_evidence = serialized.get("provider_evidence_summary") or {}
    rollout_summary = serialized.get("rollout_policy_summary") or {}

    sheets = [
        {
            "title": "Filing Summary",
            "headers": ["Field", "Value"],
            "rows": [
                ["Filing ID", serialized.get("id") or ""],
                ["Provider", serialized.get("provider") or ""],
                ["Return Type", serialized.get("return_type") or ""],
                ["Status", serialized.get("status") or ""],
                ["Workspace", serialized.get("workspace_name") or ""],
                ["Client", serialized.get("client_name") or ""],
                ["GSTIN", serialized.get("gstin_value") or ""],
                ["Compliance Period", serialized.get("compliance_period_label") or ""],
                ["Provider Reference", serialized.get("provider_reference_id") or ""],
                ["ARN", serialized.get("arn") or ""],
                ["Submitted At", serialized.get("submitted_at") or ""],
                ["Filed At", serialized.get("filed_at") or ""],
                ["Approved By", serialized.get("approved_by_name") or ""],
                ["Filed By", serialized.get("filed_by_name") or ""],
            ],
        },
        {
            "title": "Support Summary",
            "headers": ["Field", "Value"],
            "rows": [
                ["Recommended Action", support_status.get("recommended_action") or ""],
                ["Summary Reason", support_status.get("summary_reason") or ""],
                ["Latest Message", support_status.get("latest_message") or ""],
                ["Intervention Count", support_status.get("intervention_count") or 0],
                ["Can Retry", serialized.get("recovery_actions", {}).get("can_retry")],
                ["Can Resync", serialized.get("recovery_actions", {}).get("can_resync")],
                ["Provider Stage", support_status.get("provider_stage") or ""],
            ],
        },
        {
            "title": "Rollout Policy",
            "headers": ["Field", "Value"],
            "rows": [
                ["Enforced", rollout_summary.get("enforced")],
                ["Policy Present", rollout_summary.get("policy_present")],
                ["Policy Scope", ", ".join(rollout_summary.get("policy_scope") or [])],
                ["Live Submission Allowed", rollout_summary.get("live_submission_allowed")],
                ["Live Status Sync Allowed", rollout_summary.get("live_status_sync_allowed")],
                ["Submission Reason", rollout_summary.get("submission_reason") or ""],
                ["Status Sync Reason", rollout_summary.get("status_sync_reason") or ""],
                ["Notes", rollout_summary.get("notes") or ""],
            ],
        },
        {
            "title": "Latest Attempt",
            "headers": ["Field", "Value"],
            "rows": [
                ["Attempt Number", latest_attempt.get("attempt_number") or ""],
                ["Status", latest_attempt.get("status") or ""],
                ["Provider Request ID", latest_attempt.get("provider_request_id") or ""],
                ["Failure Code", latest_attempt.get("failure_code") or ""],
                ["Failure Message", latest_attempt.get("failure_message") or ""],
                ["Started At", latest_attempt.get("started_at") or ""],
                ["Submitted At", latest_attempt.get("submitted_at") or ""],
                ["Completed At", latest_attempt.get("completed_at") or ""],
            ],
        },
        {
            "title": "Provider Evidence",
            "headers": ["Field", "Value"],
            "rows": [
                ["Provider Stage", provider_evidence.get("provider_stage") or ""],
                ["Latest Message", provider_evidence.get("latest_message") or ""],
                ["Next Action", provider_evidence.get("next_action") or ""],
                ["Auth Session ID", provider_evidence.get("auth_session_id") or ""],
                ["Ops Requested", ", ".join(provider_evidence.get("operations_requested") or [])],
                ["Ops Completed", ", ".join(provider_evidence.get("operations_completed") or [])],
                ["Ops Failed", ", ".join(provider_evidence.get("operations_failed") or [])],
                ["Evidence Stored", ", ".join([key for key, value in (provider_evidence.get("evidence_available") or {}).items() if value])],
                ["Latest Failure Code", (provider_evidence.get("latest_failure") or {}).get("code") or ""],
                ["Latest Failure Message", (provider_evidence.get("latest_failure") or {}).get("message") or ""],
            ],
        },
        {
            "title": "Support Actions",
            "headers": ["Action", "Allowed", "Reason"],
            "rows": [
                [
                    action.get("label") or action.get("action") or "",
                    action.get("allowed"),
                    action.get("reason") or "",
                ]
                for action in support_actions.get("actions", [])
            ],
        },
        {
            "title": "Operational Alerts",
            "headers": ["Code", "Severity", "Title", "Message"],
            "rows": [
                [
                    alert.get("code") or "",
                    alert.get("severity") or "",
                    alert.get("title") or "",
                    alert.get("message") or "",
                ]
                for alert in serialized.get("operational_alerts", [])
            ],
        },
        {
            "title": "Incident Notes",
            "headers": ["Created At", "Title", "Severity", "Status", "Alert Code", "Note", "Resolved By", "Resolved At"],
            "rows": [
                [
                    note.get("created_at") or "",
                    note.get("title") or "",
                    note.get("severity") or "",
                    note.get("status") or "",
                    note.get("alert_code") or "",
                    note.get("note") or "",
                    note.get("resolved_by_name") or "",
                    note.get("resolved_at") or "",
                ]
                for note in serialized.get("incident_notes", [])
            ],
        },
        {
            "title": "Interventions",
            "headers": ["Created At", "Label", "New Status", "Actor", "Note"],
            "rows": [
                [
                    event.get("created_at") or "",
                    event.get("label") or "",
                    event.get("new_status") or "",
                    event.get("actor_name") or "",
                    event.get("note") or "",
                ]
                for event in serialized.get("intervention_history", [])
            ],
        },
        {
            "title": "Audit Trail",
            "headers": ["Date", "Actor", "Action", "Entity Type", "Entity ID", "Metadata"],
            "rows": (
                [
                    audit_log.created_at.isoformat() if audit_log.created_at else "",
                    display_user(audit_log.actor),
                    audit_log.action,
                    audit_log.entity_type,
                    str(audit_log.entity_id),
                    summarize_json(audit_log.metadata),
                ]
                for audit_log in AuditLog.objects.filter(entity_type="ReturnFiling", entity_id=filing.id, is_active=True)
                .select_related("actor")
                .order_by("-created_at")[:250]
            ),
        },
    ]
    return build_multi_sheet_xlsx_response(
        sheets=sheets,
        filename=f"filing-evidence-pack-{filing.id}.xlsx",
    )




def export_audit_logs(queryset: QuerySet[AuditLog]) -> HttpResponse:
    rows = (
        [
            audit_log.created_at.isoformat() if audit_log.created_at else "",
            display_user(audit_log.actor),
            audit_log.action,
            audit_log.entity_type,
            str(audit_log.entity_id),
            str(audit_log.client_id_ref or ""),
            str(audit_log.gstin_id_ref or ""),
            str(audit_log.compliance_period_id_ref or ""),
            summarize_json(audit_log.before_state),
            summarize_json(audit_log.after_state),
        ]
        for audit_log in queryset.iterator(chunk_size=2000)
    )
    return build_xlsx_response(
        sheet_title="Audit Logs",
        headers=[
            "Date",
            "Actor",
            "Action",
            "Entity Type",
            "Entity ID",
            "Client",
            "GSTIN",
            "Period",
            "Before Summary",
            "After Summary",
        ],
        rows=rows,
        filename="audit-logs.xlsx",
    )


def export_close_manager_report(*, workspace_id: str, days: int | str = 7) -> HttpResponse:
    dashboard_summary = build_close_manager_dashboard(workspace_id=workspace_id)
    report = build_close_manager_report(workspace_id=workspace_id, days=days)
    sheets = [
        {
            "title": "Summary",
            "headers": ["Metric", "Value", "Detail"],
            "rows": [
                ["Workspace", (dashboard_summary.get("workspace") or {}).get("name") or "", "Close-manager report scope"],
                ["Window (days)", report.get("window_days") or 0, "Automation reporting window"],
                ["Open assignments", dashboard_summary.get("open_assignment_count") or 0, "Open or in-progress remediation assignments"],
                ["Escalated assignments", dashboard_summary.get("escalated_assignment_count") or 0, "Assignments escalated for manager attention"],
                ["Overdue assignments", dashboard_summary.get("overdue_assignment_count") or 0, "Assignments past due"],
                ["Stale assignments", dashboard_summary.get("stale_assignment_count") or 0, "Assignments with no recent movement"],
                ["Follow-ups due today", dashboard_summary.get("follow_ups_due_today_count") or 0, "Open follow-ups due today or overdue"],
                ["Open follow-ups", dashboard_summary.get("open_follow_up_count") or 0, "Open reminder or review actions"],
                ["Digests generated", report.get("summary", {}).get("digests_generated") or 0, "Generated close-manager digests"],
                ["Digests dispatched", report.get("summary", {}).get("digests_dispatched") or 0, "Successfully dispatched digests"],
                ["Digests acknowledged", report.get("summary", {}).get("digests_acknowledged") or 0, "Acknowledged digests"],
                ["Digest failures", report.get("summary", {}).get("digest_failures") or 0, "Digest dispatch failures"],
                ["Reminders sent", report.get("summary", {}).get("reminders_sent") or 0, "Follow-up reminders sent"],
                ["Follow-ups completed", report.get("summary", {}).get("follow_ups_completed") or 0, "Follow-ups marked completed"],
                ["Follow-ups dismissed", report.get("summary", {}).get("follow_ups_dismissed") or 0, "Follow-ups dismissed"],
                ["Auto escalations", report.get("summary", {}).get("auto_escalations") or 0, "Assignments auto-escalated by automation"],
            ],
        },
        {
            "title": "Client Period Queues",
            "headers": ["Client", "Period", "GSTIN", "Open", "In Progress", "Escalated", "Overdue", "Follow-ups Due"],
            "rows": [
                [
                    queue.get("client_name") or "",
                    queue.get("period") or "",
                    queue.get("gstin_value") or "",
                    queue.get("open_assignments") or 0,
                    queue.get("in_progress_assignments") or 0,
                    queue.get("escalated_assignments") or 0,
                    queue.get("overdue_assignments") or 0,
                    queue.get("follow_ups_due") or 0,
                ]
                for queue in dashboard_summary.get("queues", [])
            ],
        },
        {
            "title": "Owner Workload",
            "headers": ["Owner", "Open Assignments", "Overdue", "Escalated"],
            "rows": [
                [
                    entry.get("name") or "",
                    entry.get("count") or 0,
                    entry.get("overdue") or 0,
                    entry.get("escalated") or 0,
                ]
                for entry in dashboard_summary.get("owner_workload", [])
            ],
        },
        {
            "title": "Due Follow-ups",
            "headers": ["Title", "Type", "Status", "Client", "Period", "Assignment", "Assigned To", "Remind At"],
            "rows": [
                [
                    follow_up.get("title") or "",
                    follow_up.get("follow_up_type") or "",
                    follow_up.get("status") or "",
                    follow_up.get("client_name") or "",
                    follow_up.get("period") or "",
                    follow_up.get("assignment_title") or "",
                    follow_up.get("assigned_to_name") or "",
                    follow_up.get("remind_at") or "",
                ]
                for follow_up in dashboard_summary.get("next_follow_ups", [])
            ],
        },
        {
            "title": "Automation Daily",
            "headers": ["Date", "Digests Generated", "Digests Dispatched", "Digests Acknowledged", "Digest Failures", "Reminders Sent", "Follow-ups Completed", "Follow-ups Dismissed", "Auto Escalations"],
            "rows": [
                [
                    entry.get("date") or "",
                    entry.get("digests_generated") or 0,
                    entry.get("digests_dispatched") or 0,
                    entry.get("digests_acknowledged") or 0,
                    entry.get("digest_failures") or 0,
                    entry.get("reminders_sent") or 0,
                    entry.get("follow_ups_completed") or 0,
                    entry.get("follow_ups_dismissed") or 0,
                    entry.get("auto_escalations") or 0,
                ]
                for entry in report.get("daily", [])
            ],
        },
        {
            "title": "Automation Activity",
            "headers": ["When", "Actor", "Action", "Entity Type", "Entity ID", "Description"],
            "rows": [
                [
                    activity.get("created_at") or "",
                    activity.get("actor_name") or "",
                    activity.get("action") or "",
                    activity.get("entity_type") or "",
                    activity.get("entity_id") or "",
                    activity.get("description") or "",
                ]
                for activity in report.get("recent_activity", [])
            ],
        },
    ]
    return build_multi_sheet_xlsx_response(
        sheets=sheets,
        filename=f"close-manager-report-{workspace_id}.xlsx",
    )


def format_decimal(value) -> str:
    if value in (None, ""):
        return "0.00"
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    try:
        return f"{Decimal(str(value)):.2f}"
    except Exception:
        return str(value)


def summarize_json(value) -> str:
    if not value:
        return ""
    try:
        payload = json.dumps(value, ensure_ascii=True, sort_keys=True)
    except TypeError:
        payload = str(value)
    return payload[:500]


def display_user(user) -> str:
    if user is None:
        return "System"
    full_name = user.get_full_name().strip()
    return full_name or user.username


def finalize_workbook_response(*, workbook: Workbook, filename: str) -> HttpResponse:
    buffer = SpooledTemporaryFile(max_size=1024 * 1024, mode="w+b")
    workbook.save(buffer)
    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type=XLSX_CONTENT_TYPE,
    )


def append_sheet(*, worksheet, title: str, headers: list[str], rows: Iterable[Sequence[object]]) -> None:
    worksheet.title = title[:31]
    worksheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    header_cells = []
    column_widths: list[int] = []
    for index, header in enumerate(headers, start=1):
        cell = WriteOnlyCell(worksheet, value=header)
        cell.font = header_font
        header_cells.append(cell)
        width = max(18, len(header) + 2)
        column_widths.append(width)
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.append(header_cells)

    for row in rows:
        normalized_row = [normalize_excel_value(value) for value in row]
        for index, value in enumerate(normalized_row):
            if index >= len(column_widths) or value is None:
                continue
            column_widths[index] = min(48, max(column_widths[index], len(str(value)) + 2))
        worksheet.append(normalized_row)

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def normalize_excel_value(value):
    if isinstance(value, UUID):
        return str(value)
    return value
