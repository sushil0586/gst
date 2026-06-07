import re
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.gst_transactions.models import GSTTransaction
from apps.reconciliation.models import ReconciliationItem, ReconciliationRun
from apps.returns.models import ReturnPreparation

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCUMENT_REF_PATTERN = re.compile(r"^(?P<prefix>[A-Za-z\-_/]+)?(?P<number>\d+)$")


def export_gstr1_workbook(*, compliance_period, prepared_return: ReturnPreparation | None = None) -> HttpResponse:
    transactions = list(
        GSTTransaction.objects.filter(
            is_active=True,
            compliance_period=compliance_period,
            transaction_type__in=["sales", "credit_note", "debit_note", "advance_received", "advance_adjusted"],
        )
        .select_related("gstin", "client")
        .order_by("transaction_date", "reference_number")
    )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_snapshot = prepared_return.summary_snapshot if prepared_return else {}
    outward_supplies = summary_snapshot.get("outward_supplies", {}) if isinstance(summary_snapshot, dict) else {}
    period_exceptions = summary_snapshot.get("period_exceptions", {}) if isinstance(summary_snapshot, dict) else {}
    client = compliance_period.gstin.client
    gstin = compliance_period.gstin

    sales_transactions = [transaction for transaction in transactions if transaction.transaction_type == "sales"]
    amendment_transactions = [transaction for transaction in transactions if is_amendment_transaction(transaction)]
    non_amendment_sales_transactions = [transaction for transaction in sales_transactions if not is_amendment_transaction(transaction)]
    export_transactions = [transaction for transaction in non_amendment_sales_transactions if special_supply_type(transaction)]
    regular_sales_transactions = [transaction for transaction in non_amendment_sales_transactions if not special_supply_type(transaction)]
    ecommerce_transactions = [transaction for transaction in non_amendment_sales_transactions if ecommerce_gstin(transaction)]
    amendment_ecommerce_transactions = [transaction for transaction in amendment_transactions if ecommerce_gstin(transaction)]
    credit_debit_notes = [transaction for transaction in transactions if transaction.transaction_type in {"credit_note", "debit_note"}]
    advance_received_transactions = [transaction for transaction in transactions if transaction.transaction_type == "advance_received"]
    advance_adjusted_transactions = [transaction for transaction in transactions if transaction.transaction_type == "advance_adjusted"]
    b2b_transactions = [transaction for transaction in regular_sales_transactions if transaction.counterparty_gstin]
    b2cl_transactions = [transaction for transaction in regular_sales_transactions if not transaction.counterparty_gstin and is_large_interstate_invoice(transaction)]
    b2cs_transactions = [transaction for transaction in regular_sales_transactions if not transaction.counterparty_gstin and not is_large_interstate_invoice(transaction)]
    hsn_transactions = sales_transactions + credit_debit_notes

    append_sheet(
        worksheet=workbook.create_sheet("Section Summary"),
        title="Section Summary",
        headers=["Section", "Documents", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_section_summary_rows(
            b2b_transactions=b2b_transactions,
            b2cl_transactions=b2cl_transactions,
            b2cs_transactions=b2cs_transactions,
            credit_debit_notes=credit_debit_notes,
            advance_received_transactions=advance_received_transactions,
            advance_adjusted_transactions=advance_adjusted_transactions,
            export_transactions=export_transactions,
            amendment_transactions=amendment_transactions,
            ecommerce_transactions=ecommerce_transactions,
        ),
    )

    append_sheet(
        worksheet=workbook.create_sheet("HSN Summary"),
        title="HSN Summary",
        headers=["HSN/SAC", "Service", "GST Rate", "Qty", "Taxable", "CGST", "SGST", "IGST", "Cess", "Docs"],
        rows=build_hsn_review_summary_rows(hsn_transactions),
    )

    append_sheet(
        worksheet=workbook.create_sheet("Document Summary"),
        title="Document Summary",
        headers=["Doc Type", "Series", "Min No", "Max No", "Total", "Cancelled"],
        rows=build_document_review_summary_rows(transactions),
    )

    append_sheet(
        worksheet=workbook.create_sheet("Nil Exempt"),
        title="Nil Exempt",
        headers=["Taxability", "Taxable", "CGST", "SGST", "IGST", "Cess"],
        rows=build_nil_exempt_review_rows(sales_transactions),
    )

    append_sheet(
        worksheet=workbook.create_sheet("Validations"),
        title="Validations",
        headers=["Code", "Severity", "Message", "Invoice ID", "Invoice Number", "Field"],
        rows=build_gstr1_validation_rows(transactions),
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Period Exceptions"),
        title="Period Exceptions",
        headers=["Document Number", "Document Date", "Transaction Type", "Category", "Reason", "Selected Period"],
        rows=build_period_exception_rows(period_exceptions),
        empty_message="No out-of-period exceptions were captured for this return.",
    )

    append_sheet(
        worksheet=workbook.create_sheet("1_3_1_3 1 2 3 Taxpayer Details"),
        title="1_3_1_3 1 2 3 Taxpayer Details",
        headers=["Gstin", "Legal Name", "Trade Name", "Previous Financial Year Aggregate Turnover"],
        rows=[[
            gstin.gstin,
            client.legal_name,
            client.trade_name or client.legal_name,
            "0.00",
        ]],
    )

    append_sheet(
        worksheet=workbook.create_sheet("4 4 B2B"),
        title="4 4 B2B",
        headers=[
            "Invoice Id",
            "Invoice Number",
            "Invoice Date",
            "Customer Name",
            "Customer Gstin",
            "Place Of Supply State Code",
            "Taxability",
            "Hsn Sac Code",
            "Is Service",
            "Taxable Amount",
            "Gst Rate",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Grand Total",
            "Reverse Charge",
            "Reported Taxable Amount",
            "Reported Cgst Amount",
            "Reported Sgst Amount",
            "Reported Igst Amount",
            "Reported Cess Amount",
            "Rcm Contract",
        ],
        rows=build_gstr1_b2b_rows(b2b_transactions),
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("5 5 B2CL (Large)"),
        title="5 5 B2CL (Large)",
        headers=[
            "Invoice Id",
            "Invoice Number",
            "Invoice Date",
            "Customer Name",
            "Place Of Supply State Code",
            "Taxability",
            "Hsn Sac Code",
            "Is Service",
            "Taxable Amount",
            "Gst Rate",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Grand Total",
            "Reported Taxable Amount",
            "Reported Cgst Amount",
            "Reported Sgst Amount",
            "Reported Igst Amount",
            "Reported Cess Amount",
        ],
        rows=build_gstr1_b2cl_rows(b2cl_transactions),
        empty_message="No rows for selected scope.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("6 6 Exports Deemed Exports SEZ"),
        title="6 6 Exports Deemed Exports SEZ",
        headers=[
            "Invoice Number",
            "Invoice Date",
            "Customer Name",
            "Customer Gstin",
            "Special Supply Type",
            "Place Of Supply State Code",
            "Port Code",
            "Shipping Bill Number",
            "Shipping Bill Date",
            "Hsn Sac Code",
            "Is Service",
            "Taxable Amount",
            "Gst Rate",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Grand Total",
        ],
        rows=build_gstr1_export_rows(export_transactions),
        empty_message="No rows for selected scope.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("7 7 B2CS"),
        title="7 7 B2CS",
        headers=[
            "Place Of Supply State Code",
            "Taxability",
            "Taxability Label",
            "Gst Rate",
            "Taxable Amount",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Grand Total",
            "Ecommerce Gstin",
        ],
        rows=build_gstr1_b2cs_rows(b2cs_transactions, gstin.state_code),
        empty_message="No rows for selected scope.",
    )

    append_sheet(
        worksheet=workbook.create_sheet("8 8 Nil Rated Exempt Non-GST"),
        title="8 8 Nil Rated Exempt Non-GST",
        headers=["Description", "Nil Rated Supplies", "Exempted Supplies", "Non-GST Supplies"],
        rows=build_gstr1_nil_exempt_rows(sales_transactions),
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("9 9 Amendments (4 5 6)"),
        title="9 9 Amendments (4 5 6)",
        headers=[
            "Target Section",
            "Current Document Number",
            "Current Document Date",
            "Original Document Number",
            "Original Document Date",
            "Original Period",
            "Original Counterparty Gstin",
            "Customer Name",
            "Ecommerce Gstin",
            "Special Supply Type",
            "Taxable Amount",
            "Tax Amount",
        ],
        rows=build_gstr1_amendment_rows(amendment_transactions),
        empty_message="No amendment rows for selected scope.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("10 10 CDNUR"),
        title="10 10 CDNUR",
        headers=[
            "Recipient Type",
            "Note Number",
            "Note Date",
            "Note Type",
            "Customer Name",
            "Customer Gstin",
            "Place Of Supply State Code",
            "Taxability",
            "Hsn Sac Code",
            "Is Service",
            "Taxable Amount",
            "Gst Rate",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Grand Total",
            "Reported Taxable Amount",
            "Reported Cgst Amount",
            "Reported Sgst Amount",
            "Reported Igst Amount",
            "Reported Cess Amount",
        ],
        rows=build_gstr1_note_rows(credit_debit_notes),
        empty_message="No rows for selected scope.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("11 11 Advances and Adjustments"),
        title="11 11 Advances and Adjustments",
        headers=["Section", "Grouped Rows", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_advance_summary_rows(
            advance_received_transactions=advance_received_transactions,
            advance_adjusted_transactions=advance_adjusted_transactions,
        ),
        empty_message="No rows for selected scope.",
    )
    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("11A Advances"),
        title="11A Advances",
        headers=[
            "Place Of Supply State Code",
            "Supply Type",
            "Gst Rate",
            "Gross Advance Amount",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Tax Amount",
            "Grand Total",
            "Document Count",
        ],
        rows=build_gstr1_advance_rows(advance_received_transactions),
        empty_message="No rows for selected scope.",
    )
    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("11B Advances"),
        title="11B Advances",
        headers=[
            "Place Of Supply State Code",
            "Supply Type",
            "Gst Rate",
            "Gross Advance Adjusted",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
            "Tax Amount",
            "Grand Total",
            "Document Count",
        ],
        rows=build_gstr1_advance_rows(advance_adjusted_transactions),
        empty_message="No rows for selected scope.",
    )

    append_sheet(
        worksheet=workbook.create_sheet("12 12 HSN Summary"),
        title="12 12 HSN Summary",
        headers=[
            "Hsn Sac Code",
            "Is Service",
            "Gst Rate",
            "Total Qty",
            "Taxable Value",
            "Cgst Amount",
            "Sgst Amount",
            "Igst Amount",
            "Cess Amount",
        ],
        rows=build_hsn_summary_rows(hsn_transactions),
    )

    append_sheet(
        worksheet=workbook.create_sheet("13 13 Documents Issued"),
        title="13 13 Documents Issued",
        headers=[
            "Doc Type",
            "Doc Type Label",
            "Doc Code",
            "Document Count",
            "Cancelled Count",
            "Min Doc No",
            "Max Doc No",
        ],
        rows=build_document_summary_rows(transactions),
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("14 14 Supplier ECO GSTIN-wise S"),
        title="14 14 Supplier ECO GSTIN-wise S",
        headers=["Ecommerce Gstin", "Section", "Place Of Supply", "Rate", "Document Count", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_ecommerce_rows(ecommerce_transactions, section_filter="table_14"),
        empty_message="No e-commerce operator supply rows for selected scope.",
    )
    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("14A 14A Amendments to Table 14"),
        title="14A 14A Amendments to Table 14",
        headers=["Ecommerce Gstin", "Section", "Place Of Supply", "Rate", "Document Count", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_ecommerce_rows(amendment_ecommerce_transactions, section_filter="table_14"),
        empty_message="No amendment rows for selected scope.",
    )
    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("15 15 ECO Operator GSTIN-wise B"),
        title="15 15 ECO Operator GSTIN-wise B",
        headers=["Ecommerce Gstin", "Section", "Place Of Supply", "Rate", "Document Count", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_ecommerce_rows(ecommerce_transactions, section_filter="table_15"),
        empty_message="No e-commerce operator rows for selected scope.",
    )
    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("15A 15A Amendments to Table 15"),
        title="15A 15A Amendments to Table 15",
        headers=["Ecommerce Gstin", "Section", "Place Of Supply", "Rate", "Document Count", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
        rows=build_gstr1_ecommerce_rows(amendment_ecommerce_transactions, section_filter="table_15"),
        empty_message="No amendment rows for selected scope.",
    )

    return workbook_response(workbook=workbook, filename=f"gstr1_{compliance_period.period}.xlsx")


def export_gstr3b_workbook(*, compliance_period, prepared_return: ReturnPreparation | None = None) -> HttpResponse:
    outward_transactions = list(
        GSTTransaction.objects.filter(
            is_active=True,
            compliance_period=compliance_period,
            transaction_type__in=["sales", "credit_note", "debit_note"],
        )
        .select_related("gstin", "client")
        .order_by("transaction_date", "reference_number")
    )
    purchase_transactions = list(
        GSTTransaction.objects.filter(
            is_active=True,
            compliance_period=compliance_period,
            transaction_type__in=["purchase", "gstr_2b"],
        ).order_by("transaction_date", "reference_number")
    )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_snapshot = prepared_return.summary_snapshot if prepared_return else {}
    outward_supplies = summary_snapshot.get("outward_supplies", {}) if isinstance(summary_snapshot, dict) else {}
    itc_summary = summary_snapshot.get("itc_summary", {}) if isinstance(summary_snapshot, dict) else {}
    reconciliation_summary = summary_snapshot.get("reconciliation", {}) if isinstance(summary_snapshot, dict) else {}
    period_exceptions = summary_snapshot.get("period_exceptions", {}) if isinstance(summary_snapshot, dict) else {}
    client = compliance_period.gstin.client
    gstin = compliance_period.gstin
    latest_run = (
        ReconciliationRun.objects.filter(
            is_active=True,
            compliance_period=compliance_period,
            run_type=ReconciliationRun.RunType.GSTR_2B_PURCHASE,
            status=ReconciliationRun.RunStatus.COMPLETED,
        )
        .order_by("-processed_at", "-created_at")
        .first()
    )
    portal_head_totals = calculate_reconciliation_portal_head_totals(latest_run)

    append_sheet(
        worksheet=workbook.create_sheet("Summary"),
        title="Summary",
        headers=["Field", "Value"],
        rows=[
            ["Return Type", "GSTR-3B"],
            ["Workspace", client.workspace.name],
            ["Client", client.legal_name],
            ["GSTIN", gstin.gstin],
            ["Period", compliance_period.period],
            ["Preparation Status", prepared_return.status if prepared_return else "not_prepared"],
            ["Prepared By", display_user(prepared_return.prepared_by) if prepared_return else "System"],
            ["Approved By", display_user(prepared_return.approved_by) if prepared_return and prepared_return.approved_by else ""],
            ["Filed By", display_user(prepared_return.filed_by) if prepared_return and prepared_return.filed_by else ""],
            ["Filed At", prepared_return.filed_at.isoformat() if prepared_return and prepared_return.filed_at else ""],
            ["ARN", prepared_return.arn if prepared_return else ""],
            ["Outward Taxable Value", outward_supplies.get("outward_taxable_value", "0.00")],
            ["Outward Tax Liability", outward_supplies.get("outward_tax_liability", "0.00")],
            ["Books ITC", itc_summary.get("books_itc", "0.00")],
            ["2B Reflected ITC", itc_summary.get("reflected_itc", "0.00")],
            ["Claim-ready ITC", itc_summary.get("claim_ready_itc", itc_summary.get("eligible_itc", "0.00"))],
            ["Pending in 2B ITC", itc_summary.get("pending_2b_itc", "0.00")],
            ["Pending Review ITC", itc_summary.get("pending_review_itc", "0.00")],
            ["Blocked ITC", itc_summary.get("blocked_itc", "0.00")],
            ["Timing Difference ITC", itc_summary.get("timing_difference_itc", "0.00")],
            ["Vendor Follow-up ITC", itc_summary.get("vendor_followup_required_itc", "0.00")],
            ["Eligible ITC", itc_summary.get("eligible_itc", "0.00")],
            ["ITC At Risk", itc_summary.get("itc_at_risk", "0.00")],
            ["Deferred / Blocked ITC", itc_summary.get("deferred_blocked_itc", "0.00")],
            ["Net Tax Payable", itc_summary.get("net_tax_payable", "0.00")],
            ["Unresolved Mismatch Count", itc_summary.get("unresolved_mismatch_count", 0)],
            ["Period Exception Count", period_exceptions.get("count", 0) if isinstance(period_exceptions, dict) else 0],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("3.1 Outward Supplies"),
        title="3.1 Outward Supplies",
        headers=[
            "Nature Of Supplies",
            "Taxable Value",
            "Integrated Tax",
            "Central Tax",
            "State/UT Tax",
            "Cess",
            "Total Tax",
        ],
        rows=build_gstr3b_outward_rows(outward_transactions),
    )

    append_sheet(
        worksheet=workbook.create_sheet("3.2 Inter-State Supplies"),
        title="3.2 Inter-State Supplies",
        headers=[
            "Nature Of Recipient",
            "Place Of Supply",
            "Taxable Value",
            "Integrated Tax",
        ],
        rows=build_gstr3b_interstate_rows(outward_transactions, gstin.state_code),
    )

    append_sheet(
        worksheet=workbook.create_sheet("4 Eligible ITC"),
        title="4 Eligible ITC",
        headers=[
            "ITC Category",
            "Integrated Tax",
            "Central Tax",
            "State/UT Tax",
            "Cess",
            "Total",
        ],
        rows=[
            [
                "All other ITC",
                format_decimal(portal_head_totals["eligible"]["igst"]),
                format_decimal(portal_head_totals["eligible"]["cgst"]),
                format_decimal(portal_head_totals["eligible"]["sgst"]),
                format_decimal(portal_head_totals["eligible"]["cess"]),
                format_decimal(sum_decimal_map(portal_head_totals["eligible"])),
            ],
            [
                "ITC reversed / at risk",
                format_decimal(portal_head_totals["at_risk"]["igst"]),
                format_decimal(portal_head_totals["at_risk"]["cgst"]),
                format_decimal(portal_head_totals["at_risk"]["sgst"]),
                format_decimal(portal_head_totals["at_risk"]["cess"]),
                format_decimal(sum_decimal_map(portal_head_totals["at_risk"])),
            ],
            [
                "Deferred / blocked ITC",
                format_decimal(portal_head_totals["deferred"]["igst"]),
                format_decimal(portal_head_totals["deferred"]["cgst"]),
                format_decimal(portal_head_totals["deferred"]["sgst"]),
                format_decimal(portal_head_totals["deferred"]["cess"]),
                format_decimal(sum_decimal_map(portal_head_totals["deferred"])),
            ],
            [
                "Net eligible ITC",
                format_decimal_decimal_like(itc_summary.get("eligible_itc", "0.00"), head="igst", heads=portal_head_totals["eligible"]),
                format_decimal_decimal_like(itc_summary.get("eligible_itc", "0.00"), head="cgst", heads=portal_head_totals["eligible"]),
                format_decimal_decimal_like(itc_summary.get("eligible_itc", "0.00"), head="sgst", heads=portal_head_totals["eligible"]),
                format_decimal_decimal_like(itc_summary.get("eligible_itc", "0.00"), head="cess", heads=portal_head_totals["eligible"]),
                format_decimal(itc_summary.get("eligible_itc", "0.00")),
            ],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("5 Exempt Supplies"),
        title="5 Exempt Supplies",
        headers=[
            "Nature Of Supplies",
            "Inter-State Supplies",
            "Intra-State Supplies",
        ],
        rows=build_gstr3b_exempt_rows(outward_transactions, gstin.state_code),
    )

    append_sheet(
        worksheet=workbook.create_sheet("5.1 Interest Late Fee"),
        title="5.1 Interest Late Fee",
        headers=[
            "Description",
            "Integrated Tax",
            "Central Tax",
            "State/UT Tax",
            "Cess",
        ],
        rows=[
            ["Interest", "0.00", "0.00", "0.00", "0.00"],
            ["Late Fee", "0.00", "0.00", "0.00", "0.00"],
        ],
    )

    output_head_totals = calculate_output_tax_heads(outward_transactions)
    append_sheet(
        worksheet=workbook.create_sheet("6 Payment Of Tax"),
        title="6 Payment Of Tax",
        headers=[
            "Description",
            "Integrated Tax",
            "Central Tax",
            "State/UT Tax",
            "Cess",
            "Total",
        ],
        rows=[
            [
                "Tax liability",
                format_decimal(output_head_totals["igst"]),
                format_decimal(output_head_totals["cgst"]),
                format_decimal(output_head_totals["sgst"]),
                format_decimal(output_head_totals["cess"]),
                format_decimal(sum_decimal_map(output_head_totals)),
            ],
            [
                "ITC utilized",
                format_decimal(portal_head_totals["eligible"]["igst"]),
                format_decimal(portal_head_totals["eligible"]["cgst"]),
                format_decimal(portal_head_totals["eligible"]["sgst"]),
                format_decimal(portal_head_totals["eligible"]["cess"]),
                format_decimal(sum_decimal_map(portal_head_totals["eligible"])),
            ],
            [
                "Net tax payable",
                format_decimal(max_decimal(output_head_totals["igst"] - portal_head_totals["eligible"]["igst"])),
                format_decimal(max_decimal(output_head_totals["cgst"] - portal_head_totals["eligible"]["cgst"])),
                format_decimal(max_decimal(output_head_totals["sgst"] - portal_head_totals["eligible"]["sgst"])),
                format_decimal(max_decimal(output_head_totals["cess"] - portal_head_totals["eligible"]["cess"])),
                format_decimal(itc_summary.get("net_tax_payable", "0.00")),
            ],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("Reconciliation Impact"),
        title="Reconciliation Impact",
        headers=["Metric", "Value"],
        rows=[
            ["Latest Reconciliation Run", str(reconciliation_summary.get("latest_run_id") or "")],
            ["Matched Count", reconciliation_summary.get("matched_count", latest_run.matched_count if latest_run else 0)],
            ["Partial Match Count", reconciliation_summary.get("partial_match_count", latest_run.partial_match_count if latest_run else 0)],
            ["Missing In Books Count", reconciliation_summary.get("missing_in_books_count", latest_run.missing_in_books_count if latest_run else 0)],
            ["Missing In Portal Count", reconciliation_summary.get("missing_in_portal_count", latest_run.missing_in_portal_count if latest_run else 0)],
            ["Duplicate Count", reconciliation_summary.get("duplicate_count", latest_run.duplicate_count if latest_run else 0)],
            ["ITC Ready Count", reconciliation_summary.get("itc_ready_count", latest_run.itc_ready_count if latest_run else 0)],
            ["Pending In 2B Count", reconciliation_summary.get("itc_pending_2b_count", latest_run.itc_pending_2b_count if latest_run else 0)],
            ["Pending Review Count", reconciliation_summary.get("itc_pending_review_count", latest_run.itc_pending_review_count if latest_run else 0)],
            ["Blocked ITC Count", reconciliation_summary.get("itc_blocked_count", latest_run.itc_blocked_count if latest_run else 0)],
            ["Timing Difference Count", reconciliation_summary.get("itc_timing_difference_count", latest_run.itc_timing_difference_count if latest_run else 0)],
            [
                "Vendor Follow-up Count",
                reconciliation_summary.get(
                    "itc_vendor_followup_required_count",
                    latest_run.itc_vendor_followup_required_count if latest_run else 0,
                ),
            ],
            ["ITC At Risk", itc_summary.get("itc_at_risk", "0.00")],
            ["Deferred / Blocked ITC", itc_summary.get("deferred_blocked_itc", "0.00")],
            ["Unresolved Mismatch Count", itc_summary.get("unresolved_mismatch_count", 0)],
        ],
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Period Exceptions"),
        title="Period Exceptions",
        headers=["Document Number", "Document Date", "Transaction Type", "Category", "Reason", "Selected Period"],
        rows=build_period_exception_rows(period_exceptions),
        empty_message="No out-of-period exceptions were captured for this return.",
    )

    append_sheet(
        worksheet=workbook.create_sheet("Source Transactions"),
        title="Source Transactions",
        headers=[
            "Transaction Type",
            "Document Type",
            "Document Number",
            "Document Date",
            "Counterparty GSTIN",
            "Counterparty Name",
            "Taxable Value",
            "IGST",
            "CGST",
            "SGST",
            "Cess",
            "Tax Amount",
            "Total Amount",
            "Place Of Supply",
            "Reverse Charge",
            "Status",
        ],
        rows=[
            [
                transaction.transaction_type,
                transaction.document_type,
                transaction.reference_number,
                transaction.transaction_date.isoformat() if transaction.transaction_date else "",
                transaction.counterparty_gstin,
                transaction.counterparty_name,
                format_decimal(transaction.taxable_value),
                format_decimal(transaction.igst_amount),
                format_decimal(transaction.cgst_amount),
                format_decimal(transaction.sgst_amount),
                format_decimal(transaction.cess_amount),
                format_decimal(transaction.tax_amount),
                format_decimal(transaction.total_amount),
                transaction.place_of_supply or "",
                "Y" if transaction.reverse_charge else "N",
                transaction.status,
            ]
            for transaction in outward_transactions + purchase_transactions
        ],
    )

    return workbook_response(workbook=workbook, filename=f"gstr3b_{compliance_period.period}.xlsx")


def export_gstr9_workbook(*, compliance_period, prepared_return: ReturnPreparation | None = None) -> HttpResponse:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_snapshot = prepared_return.summary_snapshot if prepared_return else {}
    summary_snapshot = summary_snapshot if isinstance(summary_snapshot, dict) else {}
    client = compliance_period.gstin.client
    gstin = compliance_period.gstin

    outward_summary = summary_snapshot.get("outward_summary", {}) if isinstance(summary_snapshot.get("outward_summary"), dict) else {}
    itc_summary = summary_snapshot.get("itc_summary", {}) if isinstance(summary_snapshot.get("itc_summary"), dict) else {}
    liability_summary = summary_snapshot.get("liability_summary", {}) if isinstance(summary_snapshot.get("liability_summary"), dict) else {}
    annual_sections = summary_snapshot.get("annual_sections", {}) if isinstance(summary_snapshot.get("annual_sections"), dict) else {}
    source_months = summary_snapshot.get("source_months", {}) if isinstance(summary_snapshot.get("source_months"), dict) else {}
    warnings_summary = summary_snapshot.get("warnings_summary", {}) if isinstance(summary_snapshot.get("warnings_summary"), dict) else {}
    source_trace = summary_snapshot.get("source_trace", {}) if isinstance(summary_snapshot.get("source_trace"), dict) else {}

    gstr1_return_ids = [str(item) for item in source_trace.get("gstr1_return_ids", []) if item]
    gstr3b_return_ids = [str(item) for item in source_trace.get("gstr3b_return_ids", []) if item]
    linked_returns = list(
        ReturnPreparation.objects.filter(id__in=[*gstr1_return_ids, *gstr3b_return_ids])
        .select_related("compliance_period")
        .order_by("compliance_period__period", "return_type")
    )

    append_sheet(
        worksheet=workbook.create_sheet("Summary"),
        title="Summary",
        headers=["Field", "Value"],
        rows=[
            ["Return Type", "GSTR-9"],
            ["Workspace", client.workspace.name],
            ["Client", client.legal_name],
            ["GSTIN", gstin.gstin],
            ["Anchor Period", compliance_period.period],
            ["Financial Year", summary_snapshot.get("financial_year", "")],
            ["Preparation Status", prepared_return.status if prepared_return else "not_prepared"],
            ["Prepared By", display_user(prepared_return.prepared_by) if prepared_return else "System"],
            ["Approved By", display_user(prepared_return.approved_by) if prepared_return and prepared_return.approved_by else ""],
            ["Filed By", display_user(prepared_return.filed_by) if prepared_return and prepared_return.filed_by else ""],
            ["Filed At", prepared_return.filed_at.isoformat() if prepared_return and prepared_return.filed_at else ""],
            ["ARN", prepared_return.arn if prepared_return else ""],
            ["Annual Taxable Value", outward_summary.get("annual_taxable_value", "0.00")],
            ["Annual Tax Liability", outward_summary.get("annual_tax_liability", "0.00")],
            ["Claim-ready ITC", itc_summary.get("claim_ready_itc", "0.00")],
            ["ITC At Risk", itc_summary.get("itc_at_risk", "0.00")],
            ["Net Tax Payable", liability_summary.get("net_tax_payable", "0.00")],
            ["Warning Count", warnings_summary.get("warning_count", 0)],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("Annual Outward"),
        title="Annual Outward",
        headers=["Metric", "Value"],
        rows=[
            ["GSTR-1 Taxable Value", outward_summary.get("gstr1_taxable_value", "0.00")],
            ["GSTR-1 Tax Amount", outward_summary.get("gstr1_tax_amount", "0.00")],
            ["GSTR-3B Outward Taxable Value", outward_summary.get("gstr3b_outward_taxable_value", "0.00")],
            ["GSTR-3B Tax Liability", outward_summary.get("gstr3b_outward_tax_liability", "0.00")],
            ["Annual Taxable Value", outward_summary.get("annual_taxable_value", "0.00")],
            ["Annual Tax Liability", outward_summary.get("annual_tax_liability", "0.00")],
            ["Amendment Document Count", (annual_sections.get("notes_and_amendments") or {}).get("amendment_document_count", 0)],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("Annual ITC"),
        title="Annual ITC",
        headers=["Metric", "Value"],
        rows=[
            ["Books ITC", itc_summary.get("books_itc", "0.00")],
            ["2B Reflected ITC", itc_summary.get("reflected_itc", "0.00")],
            ["Claim-ready ITC", itc_summary.get("claim_ready_itc", "0.00")],
            ["Pending in 2B ITC", itc_summary.get("pending_2b_itc", "0.00")],
            ["Pending Review ITC", itc_summary.get("pending_review_itc", "0.00")],
            ["Blocked ITC", itc_summary.get("blocked_itc", "0.00")],
            ["Timing Difference ITC", itc_summary.get("timing_difference_itc", "0.00")],
            ["Vendor Follow-up ITC", itc_summary.get("vendor_followup_required_itc", "0.00")],
            ["ITC At Risk", itc_summary.get("itc_at_risk", "0.00")],
            ["Annual Claim-ready ITC", liability_summary.get("annual_claim_ready_itc", "0.00")],
            ["Net Tax Payable", liability_summary.get("net_tax_payable", "0.00")],
        ],
    )

    append_sheet(
        worksheet=workbook.create_sheet("Source Months"),
        title="Source Months",
        headers=["Category", "Periods"],
        rows=[
            ["Expected", ", ".join(str(item) for item in source_months.get("expected_periods", []))],
            ["Available", ", ".join(str(item) for item in source_months.get("available_periods", []))],
            ["Missing", ", ".join(str(item) for item in source_months.get("missing_periods", []))],
            ["GSTR-1 Prepared", ", ".join(str(item) for item in source_months.get("gstr1_prepared_periods", []))],
            ["GSTR-3B Prepared", ", ".join(str(item) for item in source_months.get("gstr3b_prepared_periods", []))],
            ["Blocked", ", ".join(str(item) for item in source_months.get("blocked_source_periods", []))],
            ["Failed", ", ".join(str(item) for item in source_months.get("failed_source_periods", []))],
            ["Filed", ", ".join(str(item) for item in source_months.get("filed_source_periods", []))],
        ],
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Linked Source Returns"),
        title="Linked Source Returns",
        headers=["Period", "Return Type", "Status", "Return ID"],
        rows=[
            [item.compliance_period.period, item.return_type.upper(), item.status, str(item.id)]
            for item in linked_returns
        ],
        empty_message="No monthly source returns are linked to this annual draft yet.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Warnings"),
        title="Warnings",
        headers=["Code", "Severity", "Title", "Detail"],
        rows=[
            [
                str(item.get("code", "")),
                str(item.get("severity", "")),
                str(item.get("title", "")),
                str(item.get("detail", "")),
            ]
            for item in warnings_summary.get("items", [])
            if isinstance(item, dict)
        ],
        empty_message="No annual warnings were captured for this return.",
    )

    append_sheet(
        worksheet=workbook.create_sheet("Source Exceptions"),
        title="Source Exceptions",
        headers=["Metric", "Value"],
        rows=[
            ["Period Exception Count", (annual_sections.get("source_exceptions") or {}).get("period_exception_count", 0)],
            ["Missing Month Count", (annual_sections.get("source_exceptions") or {}).get("missing_month_count", 0)],
            ["Blocked Source Count", (annual_sections.get("source_exceptions") or {}).get("blocked_source_count", 0)],
            ["Failed Source Count", (annual_sections.get("source_exceptions") or {}).get("failed_source_count", 0)],
            ["Unresolved Mismatch Count", (annual_sections.get("source_exceptions") or {}).get("unresolved_mismatch_count", 0)],
            ["Manual Review Decision Count", (annual_sections.get("source_exceptions") or {}).get("manual_review_decision_count", 0)],
        ],
    )

    return workbook_response(workbook=workbook, filename=f"gstr9_{compliance_period.period}.xlsx")


def export_gstr7_workbook(*, compliance_period, prepared_return: ReturnPreparation | None = None) -> HttpResponse:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_snapshot = prepared_return.summary_snapshot if prepared_return else {}
    summary_snapshot = summary_snapshot if isinstance(summary_snapshot, dict) else {}
    client = compliance_period.gstin.client
    gstin = compliance_period.gstin
    transactions = list(
        GSTTransaction.objects.filter(
            is_active=True,
            compliance_period=compliance_period,
            transaction_type="tds_deducted",
        )
        .select_related("gstin", "client")
        .order_by("transaction_date", "reference_number")
    )

    tds_summary = summary_snapshot.get("tds_summary", {}) if isinstance(summary_snapshot.get("tds_summary"), dict) else {}
    deductees = summary_snapshot.get("deductees", {}) if isinstance(summary_snapshot.get("deductees"), dict) else {}
    period_exceptions = summary_snapshot.get("period_exceptions", {}) if isinstance(summary_snapshot.get("period_exceptions"), dict) else {}

    append_sheet(
        worksheet=workbook.create_sheet("Summary"),
        title="Summary",
        headers=["Field", "Value"],
        rows=[
            ["Return Type", "GSTR-7"],
            ["Workspace", client.workspace.name],
            ["Client", client.legal_name],
            ["GSTIN", gstin.gstin],
            ["Period", compliance_period.period],
            ["Preparation Status", prepared_return.status if prepared_return else "not_prepared"],
            ["Prepared By", display_user(prepared_return.prepared_by) if prepared_return else "System"],
            ["Deductee Count", tds_summary.get("deductee_count", 0)],
            ["Document Count", tds_summary.get("document_count", 0)],
            ["Payment Amount", tds_summary.get("payment_amount", "0.00")],
            ["Taxable Value", tds_summary.get("taxable_value", "0.00")],
            ["IGST Deducted", tds_summary.get("igst_amount", "0.00")],
            ["CGST Deducted", tds_summary.get("cgst_amount", "0.00")],
            ["SGST Deducted", tds_summary.get("sgst_amount", "0.00")],
            ["Total TDS Deducted", tds_summary.get("tds_amount", "0.00")],
        ],
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Deductees"),
        title="Deductees",
        headers=[
            "Deductee GSTIN",
            "Deductee Name",
            "Document Count",
            "Payment Amount",
            "Taxable Value",
            "IGST Deducted",
            "CGST Deducted",
            "SGST Deducted",
            "Total TDS Deducted",
        ],
        rows=[
            [
                row.get("deductee_gstin", ""),
                row.get("deductee_name", ""),
                row.get("document_count", 0),
                row.get("payment_amount", "0.00"),
                row.get("taxable_value", "0.00"),
                row.get("igst_amount", "0.00"),
                row.get("cgst_amount", "0.00"),
                row.get("sgst_amount", "0.00"),
                row.get("tds_amount", "0.00"),
            ]
            for row in deductees.get("rows", [])
            if isinstance(row, dict)
        ],
        empty_message="No deductee rows were captured for this return.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Source Rows"),
        title="Source Rows",
        headers=[
            "Document Number",
            "Document Date",
            "Deductee GSTIN",
            "Deductee Name",
            "Payment Amount",
            "Taxable Value",
            "IGST Deducted",
            "CGST Deducted",
            "SGST Deducted",
            "Total TDS Deducted",
        ],
        rows=[
            [
                transaction.reference_number,
                transaction.transaction_date.isoformat() if transaction.transaction_date else "",
                transaction.counterparty_gstin,
                transaction.counterparty_name,
                format_decimal(transaction.total_amount),
                format_decimal(transaction.taxable_value),
                format_decimal(transaction.igst_amount),
                format_decimal(transaction.cgst_amount),
                format_decimal(transaction.sgst_amount),
                format_decimal(transaction.tax_amount),
            ]
            for transaction in transactions
        ],
        empty_message="No TDS deducted source rows are available for the selected period.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Validations"),
        title="Validations",
        headers=["Code", "Severity", "Message", "Row", "Document Number", "Field"],
        rows=build_gstr7_validation_rows(transactions),
        empty_message="No workbook-level validation signals were generated for this return.",
    )

    append_data_or_info_sheet(
        worksheet=workbook.create_sheet("Period Exceptions"),
        title="Period Exceptions",
        headers=["Document Number", "Document Date", "Transaction Type", "Category", "Reason", "Selected Period"],
        rows=build_period_exception_rows(period_exceptions),
        empty_message="No out-of-period exceptions were captured for this return.",
    )

    return workbook_response(workbook=workbook, filename=f"gstr7_{compliance_period.period}.xlsx")


def append_sheet(*, worksheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    worksheet.title = title
    worksheet.append(headers)
    worksheet.freeze_panes = "A2"

    header_font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.font = header_font
        worksheet.column_dimensions[get_column_letter(index)].width = max(18, len(header) + 2)

    for row in rows:
        worksheet.append([normalize_excel_value(value) for value in row])

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            current_width = worksheet.column_dimensions[cell.column_letter].width or 18
            worksheet.column_dimensions[cell.column_letter].width = min(48, max(current_width, len(str(cell.value)) + 2))


def append_info_sheet(*, worksheet, title: str, message: str) -> None:
    append_sheet(worksheet=worksheet, title=title, headers=["Info"], rows=[[message]])


def append_data_or_info_sheet(*, worksheet, title: str, headers: list[str], rows: list[list[object]], empty_message: str) -> None:
    if rows:
        append_sheet(worksheet=worksheet, title=title, headers=headers, rows=rows)
        return
    append_info_sheet(worksheet=worksheet, title=title, message=empty_message)


def build_period_exception_rows(period_exceptions) -> list[list[object]]:
    if not isinstance(period_exceptions, dict):
        return []
    documents = period_exceptions.get("documents")
    if not isinstance(documents, list):
        return []
    rows = []
    for item in documents:
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                item.get("document_number", ""),
                item.get("document_date", ""),
                item.get("transaction_type", ""),
                str(item.get("category", "general")).replace("_", " "),
                item.get("reason", ""),
                item.get("selected_period", ""),
            ]
        )
    return rows


def workbook_response(*, workbook: Workbook, filename: str) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def format_decimal(value) -> str:
    if value in (None, ""):
        return "0.00"
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    try:
        return f"{Decimal(str(value)):.2f}"
    except Exception:
        return str(value)


def decimal_or_zero(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def display_user(user) -> str:
    if user is None:
        return "System"
    full_name = user.get_full_name().strip()
    return full_name or user.username


def normalize_excel_value(value):
    if isinstance(value, UUID):
        return str(value)
    return value


def date_to_iso(value) -> str:
    if value in (None, ""):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def calculate_rate(transaction: GSTTransaction) -> Decimal:
    if not transaction.taxable_value:
        return Decimal("0.00")
    return (transaction.tax_amount / transaction.taxable_value) * Decimal("100.00")


def iter_transaction_components(transaction: GSTTransaction) -> list[dict[str, object]]:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    raw_components = metadata.get("line_items")
    if not isinstance(raw_components, list) or not raw_components:
        return [
            {
                "hsn_code": hsn_code(transaction),
                "uqc": str(metadata.get("uqc") or ""),
                "quantity": decimal_or_zero(metadata.get("quantity")),
                "taxable_value": decimal_or_zero(transaction.taxable_value),
                "cgst_amount": decimal_or_zero(transaction.cgst_amount),
                "sgst_amount": decimal_or_zero(transaction.sgst_amount),
                "igst_amount": decimal_or_zero(transaction.igst_amount),
                "cess_amount": decimal_or_zero(transaction.cess_amount),
                "total_amount": decimal_or_zero(transaction.total_amount),
                "is_service": is_service_transaction(transaction),
                "description": str(metadata.get("description") or transaction.counterparty_name or "Outward supply"),
                "rate": decimal_or_zero(calculate_rate(transaction)),
                "reference_number": transaction.reference_number,
            }
        ]

    components: list[dict[str, object]] = []
    for item in raw_components:
        if not isinstance(item, dict):
            continue
        taxable_value = decimal_or_zero(item.get("taxable_value"))
        cgst_amount = decimal_or_zero(item.get("cgst_amount"))
        sgst_amount = decimal_or_zero(item.get("sgst_amount"))
        igst_amount = decimal_or_zero(item.get("igst_amount"))
        cess_amount = decimal_or_zero(item.get("cess_amount"))
        tax_amount = cgst_amount + sgst_amount + igst_amount + cess_amount
        rate = decimal_or_zero(item.get("rate"))
        if rate == Decimal("0.00") and taxable_value > Decimal("0.00"):
            rate = (tax_amount / taxable_value) * Decimal("100.00") if tax_amount else Decimal("0.00")
        total_amount = decimal_or_zero(item.get("total_amount"))
        if total_amount == Decimal("0.00"):
            total_amount = taxable_value + tax_amount
        components.append(
            {
                "hsn_code": str(item.get("hsn_code") or hsn_code(transaction)),
                "uqc": str(item.get("uqc") or metadata.get("uqc") or ""),
                "quantity": decimal_or_zero(item.get("quantity")),
                "taxable_value": taxable_value,
                "cgst_amount": cgst_amount,
                "sgst_amount": sgst_amount,
                "igst_amount": igst_amount,
                "cess_amount": cess_amount,
                "total_amount": total_amount,
                "is_service": bool(item.get("is_service") if item.get("is_service") is not None else is_service_transaction(transaction)),
                "description": str(item.get("description") or metadata.get("description") or transaction.counterparty_name or "Outward supply"),
                "rate": rate,
                "reference_number": transaction.reference_number,
            }
        )

    return components or [
        {
            "hsn_code": hsn_code(transaction),
            "uqc": str(metadata.get("uqc") or ""),
            "quantity": decimal_or_zero(metadata.get("quantity")),
            "taxable_value": decimal_or_zero(transaction.taxable_value),
            "cgst_amount": decimal_or_zero(transaction.cgst_amount),
            "sgst_amount": decimal_or_zero(transaction.sgst_amount),
            "igst_amount": decimal_or_zero(transaction.igst_amount),
            "cess_amount": decimal_or_zero(transaction.cess_amount),
            "total_amount": decimal_or_zero(transaction.total_amount),
            "is_service": is_service_transaction(transaction),
            "description": str(metadata.get("description") or transaction.counterparty_name or "Outward supply"),
            "rate": decimal_or_zero(calculate_rate(transaction)),
            "reference_number": transaction.reference_number,
        }
    ]


def is_large_interstate_invoice(transaction: GSTTransaction) -> bool:
    gstin_state = getattr(transaction.gstin, "state_code", "")
    place_of_supply = str(transaction.place_of_supply or "").strip()
    is_interstate = bool(place_of_supply and gstin_state and place_of_supply != gstin_state)
    return is_interstate and transaction.total_amount >= Decimal("250000.00")


def build_hsn_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[str, dict[str, Decimal | str]] = {}
    for transaction in transactions:
        for component in iter_transaction_components(transaction):
            hsn = str(component["hsn_code"] or "UNSPECIFIED")
            entry = grouped.setdefault(
                hsn,
                {
                    "description": str(component["description"]),
                    "uqc": str(component["uqc"]),
                    "quantity": Decimal("0.00"),
                    "total_value": Decimal("0.00"),
                    "taxable_value": Decimal("0.00"),
                    "igst_amount": Decimal("0.00"),
                    "cgst_amount": Decimal("0.00"),
                    "sgst_amount": Decimal("0.00"),
                    "cess_amount": Decimal("0.00"),
                },
            )
            entry["quantity"] += decimal_or_zero(component["quantity"])
            entry["total_value"] += decimal_or_zero(component["total_amount"])
            entry["taxable_value"] += decimal_or_zero(component["taxable_value"])
            entry["igst_amount"] += decimal_or_zero(component["igst_amount"])
            entry["cgst_amount"] += decimal_or_zero(component["cgst_amount"])
            entry["sgst_amount"] += decimal_or_zero(component["sgst_amount"])
            entry["cess_amount"] += decimal_or_zero(component["cess_amount"])

    rows = []
    for hsn, entry in sorted(grouped.items()):
        rows.append(
            [
                hsn,
                entry["description"],
                entry["uqc"],
                format_decimal(entry["quantity"]),
                format_decimal(entry["total_value"]),
                format_decimal(entry["taxable_value"]),
                format_decimal(entry["igst_amount"]),
                format_decimal(entry["cgst_amount"]),
                format_decimal(entry["sgst_amount"]),
                format_decimal(entry["cess_amount"]),
            ]
        )
    return rows


def build_hsn_summary_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[tuple[str, str, str], dict[str, Decimal | bool]] = {}
    for transaction in transactions:
        for component in iter_transaction_components(transaction):
            hsn = str(component["hsn_code"] or "UNSPECIFIED")
            is_service = bool(component["is_service"])
            rate = format_decimal(component["rate"])
            key = (hsn, "true" if is_service else "false", rate)
            entry = grouped.setdefault(
                key,
                {
                    "quantity": Decimal("0.00"),
                    "taxable_value": Decimal("0.00"),
                    "cgst_amount": Decimal("0.00"),
                    "sgst_amount": Decimal("0.00"),
                    "igst_amount": Decimal("0.00"),
                    "cess_amount": Decimal("0.00"),
                },
            )
            entry["quantity"] += decimal_or_zero(component["quantity"])
            entry["taxable_value"] += decimal_or_zero(component["taxable_value"])
            entry["cgst_amount"] += decimal_or_zero(component["cgst_amount"])
            entry["sgst_amount"] += decimal_or_zero(component["sgst_amount"])
            entry["igst_amount"] += decimal_or_zero(component["igst_amount"])
            entry["cess_amount"] += decimal_or_zero(component["cess_amount"])

    rows = []
    for (hsn, is_service, rate), entry in sorted(grouped.items()):
        rows.append(
            [
                hsn,
                is_service == "true",
                rate,
                format_decimal(entry["quantity"]),
                format_decimal(entry["taxable_value"]),
                format_decimal(entry["cgst_amount"]),
                format_decimal(entry["sgst_amount"]),
                format_decimal(entry["igst_amount"]),
                format_decimal(entry["cess_amount"]),
            ]
        )
    return rows


def build_hsn_review_summary_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[tuple[str, str, str], dict[str, Decimal | int]] = {}
    for transaction in transactions:
        seen_doc_keys: set[tuple[str, str, str]] = set()
        for component in iter_transaction_components(transaction):
            hsn = str(component["hsn_code"] or "UNSPECIFIED")
            is_service = "Y" if component["is_service"] else "N"
            rate = format_decimal(component["rate"])
            key = (hsn, is_service, rate)
            entry = grouped.setdefault(
                key,
                {
                    "qty": Decimal("0.00"),
                    "taxable": Decimal("0.00"),
                    "cgst": Decimal("0.00"),
                    "sgst": Decimal("0.00"),
                    "igst": Decimal("0.00"),
                    "cess": Decimal("0.00"),
                    "docs": 0,
                },
            )
            entry["qty"] += decimal_or_zero(component["quantity"])
            entry["taxable"] += decimal_or_zero(component["taxable_value"])
            entry["cgst"] += decimal_or_zero(component["cgst_amount"])
            entry["sgst"] += decimal_or_zero(component["sgst_amount"])
            entry["igst"] += decimal_or_zero(component["igst_amount"])
            entry["cess"] += decimal_or_zero(component["cess_amount"])
            if key not in seen_doc_keys:
                entry["docs"] += 1
                seen_doc_keys.add(key)

    rows = []
    for (hsn, is_service, rate), entry in sorted(grouped.items()):
        rows.append(
            [
                hsn,
                is_service,
                rate,
                format_decimal(entry["qty"]),
                format_decimal(entry["taxable"]),
                format_decimal(entry["cgst"]),
                format_decimal(entry["sgst"]),
                format_decimal(entry["igst"]),
                format_decimal(entry["cess"]),
                entry["docs"],
            ]
        )
    return rows


def build_document_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[str, list[str]] = {}
    for transaction in transactions:
        key = f"{transaction.transaction_type}:{transaction.document_type}"
        grouped.setdefault(key, []).append(transaction.reference_number)

    rows = []
    for key, refs in sorted(grouped.items()):
        refs = sorted(filter(None, refs))
        transaction_type, document_type = key.split(":", 1)
        rows.append(
            [
                f"{prettify_document_type(transaction_type)} - {prettify_document_type(document_type)}",
                refs[0] if refs else "",
                refs[-1] if refs else "",
                len(refs),
                0,
            ]
        )
    return rows


def build_document_review_summary_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[tuple[int, str], list[tuple[str, int | None, str]]] = {}
    for transaction in transactions:
        series, serial_number, original_reference = parse_document_reference(transaction.reference_number)
        doc_type_id = document_type_id_for_summary(transaction.transaction_type, transaction.document_type)
        doc_code = document_code_for_type(transaction.transaction_type, transaction.document_type)
        grouped.setdefault((doc_type_id, doc_code), []).append((series, serial_number, original_reference))

    rows = []
    for doc_type_id, series in sorted(grouped.keys()):
        refs = grouped[(doc_type_id, series)]
        min_value, max_value = summarize_reference_range(refs)
        rows.append(
            [
                doc_type_id,
                series,
                min_value,
                max_value,
                len(refs),
                0,
            ]
        )
    return rows


def build_document_summary_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[tuple[str, str, str], list[tuple[str, int | None, str]]] = {}
    for transaction in transactions:
        series, serial_number, original_reference = parse_document_reference(transaction.reference_number)
        key = (transaction.transaction_type, transaction.document_type, series)
        grouped.setdefault(key, []).append((series, serial_number, original_reference))

    rows = []
    for index, ((transaction_type, document_type, series), refs) in enumerate(sorted(grouped.items()), start=1):
        min_value, max_value = summarize_reference_range(refs)
        rows.append(
            [
                document_type_id_for_summary(transaction_type, document_type),
                document_type_label_for_summary(transaction_type, document_type),
                document_code_for_type(transaction_type, document_type),
                len(refs),
                0,
                min_value,
                max_value,
            ]
        )
    return rows


def prettify_document_type(document_type: str) -> str:
    return document_type.replace("_", " ").title() if document_type else "Invoice"


def document_type_id_for_summary(transaction_type: str, document_type: str) -> int:
    if transaction_type == "sales" and document_type == "invoice":
        return 1
    if transaction_type == "credit_note":
        return 5
    if transaction_type == "debit_note":
        return 6
    if transaction_type == "advance_received":
        return 7
    if transaction_type == "advance_adjusted":
        return 8
    return 99


def document_type_label_for_summary(transaction_type: str, document_type: str) -> str:
    if transaction_type == "sales" and document_type == "invoice":
        return "Tax Invoice"
    if transaction_type == "credit_note":
        return "Credit Note"
    if transaction_type == "debit_note":
        return "Debit Note"
    if transaction_type == "advance_received":
        return "Receipt Voucher"
    if transaction_type == "advance_adjusted":
        return "Advance Adjustment"
    return prettify_document_type(document_type)


def document_code_for_type(transaction_type: str, document_type: str) -> str:
    if document_type == "invoice" and transaction_type == "sales":
        return "SINV"
    if transaction_type == "credit_note":
        return "CRN"
    if transaction_type == "debit_note":
        return "DBN"
    if transaction_type == "advance_received":
        return "RCV"
    if transaction_type == "advance_adjusted":
        return "ADVADJ"
    return document_type.replace("_", "").upper()[:8] or "DOC"


def parse_document_reference(reference_number: str) -> tuple[str, int | None, str]:
    value = str(reference_number or "").strip()
    if not value:
        return "DOC", None, ""
    compact = value.replace(" ", "")
    match = DOCUMENT_REF_PATTERN.match(compact)
    if not match:
        return compact, None, value
    prefix = (match.group("prefix") or "").strip("-_/") or "DOC"
    number = int(match.group("number"))
    return prefix.upper(), number, value


def summarize_reference_range(refs: list[tuple[str, int | None, str]]) -> tuple[object, object]:
    numeric_refs = [number for _, number, _ in refs if number is not None]
    if numeric_refs:
        return min(numeric_refs), max(numeric_refs)
    original_refs = sorted(filter(None, (original for _, _, original in refs)))
    if not original_refs:
        return "", ""
    return original_refs[0], original_refs[-1]


def build_gstr1_section_summary_rows(
    *,
    b2b_transactions: list[GSTTransaction],
    b2cl_transactions: list[GSTTransaction],
    b2cs_transactions: list[GSTTransaction],
    credit_debit_notes: list[GSTTransaction],
    advance_received_transactions: list[GSTTransaction],
    advance_adjusted_transactions: list[GSTTransaction],
    export_transactions: list[GSTTransaction],
    amendment_transactions: list[GSTTransaction],
    ecommerce_transactions: list[GSTTransaction],
) -> list[list[object]]:
    return [
        ["B2B", len(b2b_transactions), *section_tax_summary(b2b_transactions)],
        ["B2CL", len(b2cl_transactions), *section_tax_summary(b2cl_transactions)],
        ["B2CS", len(b2cs_transactions), *section_tax_summary(b2cs_transactions)],
        ["CDN", len(credit_debit_notes), *section_tax_summary(credit_debit_notes)],
        ["ADV_RECEIVED", len(build_gstr1_advance_rows(advance_received_transactions)), *section_tax_summary(advance_received_transactions)],
        ["ADV_ADJUSTED", len(build_gstr1_advance_rows(advance_adjusted_transactions)), *section_tax_summary(advance_adjusted_transactions)],
        ["EXPORTS", len(build_gstr1_export_rows(export_transactions)), *section_tax_summary(export_transactions)],
        ["AMENDMENTS", len(build_gstr1_amendment_rows(amendment_transactions)), *section_tax_summary(amendment_transactions)],
        ["ECOMMERCE", len(build_gstr1_ecommerce_rows(ecommerce_transactions)), *section_tax_summary(ecommerce_transactions)],
    ]


def section_tax_summary(transactions: list[GSTTransaction]) -> list[str]:
    taxable = sum(((transaction.taxable_value or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    cgst = sum(((transaction.cgst_amount or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    sgst = sum(((transaction.sgst_amount or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    igst = sum(((transaction.igst_amount or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    cess = sum(((transaction.cess_amount or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    total = sum(((transaction.total_amount or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))
    return [
        format_decimal(taxable),
        format_decimal(cgst),
        format_decimal(sgst),
        format_decimal(igst),
        format_decimal(cess),
        format_decimal(total),
    ]


def build_gstr1_b2b_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for index, transaction in enumerate(transactions, start=1):
        components = iter_transaction_components(transaction)
        invoice_total = decimal_or_zero(transaction.total_amount)
        for component in components:
            rows.append(
                    [
                        index,
                        transaction.reference_number,
                        date_to_iso(transaction.transaction_date),
                        transaction.counterparty_name,
                    transaction.counterparty_gstin,
                    transaction.place_of_supply or "",
                    taxability_code(transaction),
                    component["hsn_code"],
                    bool(component["is_service"]),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["rate"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    invoice_total,
                    transaction.reverse_charge,
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    build_rcm_contract(transaction, table_code="TABLE_4"),
                ]
            )
    return rows


def build_gstr1_b2cl_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for index, transaction in enumerate(transactions, start=1):
        for component in iter_transaction_components(transaction):
            rows.append(
                    [
                        index,
                        transaction.reference_number,
                        date_to_iso(transaction.transaction_date),
                        transaction.counterparty_name or "Unregistered Customer",
                    transaction.place_of_supply or "",
                    taxability_code(transaction),
                    component["hsn_code"],
                    bool(component["is_service"]),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["rate"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    decimal_or_zero(transaction.total_amount),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                ]
            )
    return rows


def build_gstr1_b2cs_rows(transactions: list[GSTTransaction], home_state_code: str) -> list[list[object]]:
    rows = []
    for transaction in transactions:
        place_of_supply = str(transaction.place_of_supply or "").strip()
        taxability = taxability_code(transaction)
        taxability_label = "Intrastate" if place_of_supply == home_state_code else "Interstate"
        metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
        for component in iter_transaction_components(transaction):
            rows.append(
                [
                    place_of_supply,
                    taxability,
                    taxability_label,
                    decimal_or_zero(component["rate"]),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    decimal_or_zero(component["total_amount"]),
                    metadata.get("ecommerce_gstin", ""),
                ]
            )
    return rows


def build_gstr1_nil_exempt_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    nil_rated = Decimal("0.00")
    exempt = Decimal("0.00")
    non_gst = Decimal("0.00")
    for transaction in transactions:
        category = inferred_supply_category(transaction)
        taxable_value = transaction.taxable_value or Decimal("0.00")
        if category == "nil_rated":
            nil_rated += taxable_value
        elif category == "exempt":
            exempt += taxable_value
        elif category == "non_gst":
            non_gst += taxable_value
    if nil_rated == exempt == non_gst == Decimal("0.00"):
        return []
    return [["Outward supplies", format_decimal(nil_rated), format_decimal(exempt), format_decimal(non_gst)]]


def build_nil_exempt_review_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    totals = {
        "Nil Rated": {"taxable": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "igst": Decimal("0.00"), "cess": Decimal("0.00")},
        "Exempt": {"taxable": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "igst": Decimal("0.00"), "cess": Decimal("0.00")},
        "Non-GST": {"taxable": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "igst": Decimal("0.00"), "cess": Decimal("0.00")},
    }
    for transaction in transactions:
        category = inferred_supply_category(transaction)
        bucket = None
        if category == "nil_rated":
            bucket = "Nil Rated"
        elif category == "exempt":
            bucket = "Exempt"
        elif category == "non_gst":
            bucket = "Non-GST"
        if not bucket:
            continue
        totals[bucket]["taxable"] += transaction.taxable_value or Decimal("0.00")
        totals[bucket]["cgst"] += transaction.cgst_amount or Decimal("0.00")
        totals[bucket]["sgst"] += transaction.sgst_amount or Decimal("0.00")
        totals[bucket]["igst"] += transaction.igst_amount or Decimal("0.00")
        totals[bucket]["cess"] += transaction.cess_amount or Decimal("0.00")

    rows = []
    for label, values in totals.items():
        if all(value == Decimal("0.00") for value in values.values()):
            continue
        rows.append(
            [
                label,
                format_decimal(values["taxable"]),
                format_decimal(values["cgst"]),
                format_decimal(values["sgst"]),
                format_decimal(values["igst"]),
                format_decimal(values["cess"]),
            ]
        )
    return rows


def build_gstr1_note_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for transaction in transactions:
        for component in iter_transaction_components(transaction):
            rows.append(
                [
                    "Registered" if transaction.counterparty_gstin else "Unregistered",
                    transaction.reference_number,
                    date_to_iso(transaction.transaction_date),
                    "Credit Note" if transaction.transaction_type == "credit_note" else "Debit Note",
                    transaction.counterparty_name,
                    transaction.counterparty_gstin,
                    transaction.place_of_supply or "",
                    taxability_code(transaction),
                    component["hsn_code"],
                    bool(component["is_service"]),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["rate"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    decimal_or_zero(transaction.total_amount),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                ]
            )
    return rows


def build_gstr1_advance_summary_rows(
    *,
    advance_received_transactions: list[GSTTransaction],
    advance_adjusted_transactions: list[GSTTransaction],
) -> list[list[object]]:
    advance_received_rows = build_gstr1_advance_rows(advance_received_transactions)
    advance_adjusted_rows = build_gstr1_advance_rows(advance_adjusted_transactions)
    return [
        ["Advances Received", len(advance_received_rows), *section_tax_summary(advance_received_transactions)],
        ["Advances Adjusted", len(advance_adjusted_rows), *section_tax_summary(advance_adjusted_transactions)],
    ]


def build_gstr1_advance_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    grouped: dict[tuple[str, str, str], dict[str, Decimal | int]] = {}
    for transaction in transactions:
        place_of_supply = str(transaction.place_of_supply or "").strip() or "00"
        supply_type = "INTER" if (transaction.igst_amount or Decimal("0.00")) > Decimal("0.00") else "INTRA"
        for component in iter_transaction_components(transaction):
            rate = format_decimal(component["rate"])
            key = (place_of_supply, supply_type, rate)
            entry = grouped.setdefault(
                key,
                {
                    "taxable_value": Decimal("0.00"),
                    "cgst_amount": Decimal("0.00"),
                    "sgst_amount": Decimal("0.00"),
                    "igst_amount": Decimal("0.00"),
                    "cess_amount": Decimal("0.00"),
                    "tax_amount": Decimal("0.00"),
                    "total_amount": Decimal("0.00"),
                    "document_count": 0,
                },
            )
            entry["taxable_value"] += decimal_or_zero(component["taxable_value"])
            entry["cgst_amount"] += decimal_or_zero(component["cgst_amount"])
            entry["sgst_amount"] += decimal_or_zero(component["sgst_amount"])
            entry["igst_amount"] += decimal_or_zero(component["igst_amount"])
            entry["cess_amount"] += decimal_or_zero(component["cess_amount"])
            entry["tax_amount"] += (
                decimal_or_zero(component["cgst_amount"])
                + decimal_or_zero(component["sgst_amount"])
                + decimal_or_zero(component["igst_amount"])
                + decimal_or_zero(component["cess_amount"])
            )
            entry["total_amount"] += decimal_or_zero(component["total_amount"])
            entry["document_count"] += 1

    rows = []
    for (place_of_supply, supply_type, rate), entry in sorted(grouped.items()):
        rows.append(
            [
                place_of_supply,
                supply_type,
                decimal_or_zero(rate),
                decimal_or_zero(entry["taxable_value"]),
                decimal_or_zero(entry["cgst_amount"]),
                decimal_or_zero(entry["sgst_amount"]),
                decimal_or_zero(entry["igst_amount"]),
                decimal_or_zero(entry["cess_amount"]),
                decimal_or_zero(entry["tax_amount"]),
                decimal_or_zero(entry["total_amount"]),
                int(entry["document_count"]),
            ]
        )
    return rows


def build_gstr1_export_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for transaction in transactions:
        metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
        for component in iter_transaction_components(transaction):
            rows.append(
                [
                    transaction.reference_number,
                    date_to_iso(transaction.transaction_date),
                    transaction.counterparty_name,
                    transaction.counterparty_gstin,
                    special_supply_type_label(transaction),
                    transaction.place_of_supply or "",
                    metadata.get("port_code", ""),
                    metadata.get("shipping_bill_number", ""),
                    metadata.get("shipping_bill_date", ""),
                    component["hsn_code"],
                    bool(component["is_service"]),
                    decimal_or_zero(component["taxable_value"]),
                    decimal_or_zero(component["rate"]),
                    decimal_or_zero(component["cgst_amount"]),
                    decimal_or_zero(component["sgst_amount"]),
                    decimal_or_zero(component["igst_amount"]),
                    decimal_or_zero(component["cess_amount"]),
                    decimal_or_zero(component["total_amount"]),
                ]
            )
    return rows


def build_gstr1_amendment_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for transaction in transactions:
        rows.append(
            [
                amendment_target_section(transaction),
                transaction.reference_number,
                date_to_iso(transaction.transaction_date),
                original_document_number(transaction),
                original_document_date(transaction),
                original_period(transaction),
                original_counterparty_gstin(transaction),
                transaction.counterparty_name,
                ecommerce_gstin(transaction),
                special_supply_type_label(transaction) if special_supply_type(transaction) else "",
                decimal_or_zero(transaction.taxable_value),
                decimal_or_zero(transaction.tax_amount),
            ]
        )
    return rows


def build_gstr1_ecommerce_rows(transactions: list[GSTTransaction], section_filter: str | None = None) -> list[list[object]]:
    grouped: dict[tuple[str, str, str, str], dict[str, Decimal | int]] = {}
    for transaction in transactions:
        etin = ecommerce_gstin(transaction)
        section_code = ecommerce_section(transaction)
        if not etin or (section_filter and section_code != section_filter):
            continue
        place_of_supply = str(transaction.place_of_supply or "").strip() or "00"
        for component in iter_transaction_components(transaction):
            rate = format_decimal(component["rate"])
            key = (etin, section_code, place_of_supply, rate)
            entry = grouped.setdefault(
                key,
                {
                    "taxable_value": Decimal("0.00"),
                    "cgst_amount": Decimal("0.00"),
                    "sgst_amount": Decimal("0.00"),
                    "igst_amount": Decimal("0.00"),
                    "cess_amount": Decimal("0.00"),
                    "total_amount": Decimal("0.00"),
                    "document_count": 0,
                },
            )
            entry["taxable_value"] += decimal_or_zero(component["taxable_value"])
            entry["cgst_amount"] += decimal_or_zero(component["cgst_amount"])
            entry["sgst_amount"] += decimal_or_zero(component["sgst_amount"])
            entry["igst_amount"] += decimal_or_zero(component["igst_amount"])
            entry["cess_amount"] += decimal_or_zero(component["cess_amount"])
            entry["total_amount"] += decimal_or_zero(component["total_amount"])
            entry["document_count"] += 1

    rows = []
    for (etin, section_code, place_of_supply, rate), entry in sorted(grouped.items()):
        rows.append(
            [
                etin,
                section_code,
                place_of_supply,
                decimal_or_zero(rate),
                int(entry["document_count"]),
                decimal_or_zero(entry["taxable_value"]),
                decimal_or_zero(entry["cgst_amount"]),
                decimal_or_zero(entry["sgst_amount"]),
                decimal_or_zero(entry["igst_amount"]),
                decimal_or_zero(entry["cess_amount"]),
                decimal_or_zero(entry["total_amount"]),
            ]
        )
    return rows


def build_gstr1_validation_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    for index, transaction in enumerate(transactions, start=1):
        metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
        if transaction.transaction_type not in {"advance_received", "advance_adjusted"} and not metadata.get("hsn_code"):
            rows.append(["HSN_MISSING", "warning", "HSN/SAC is not available; workbook uses UNSPECIFIED.", index, transaction.reference_number, "hsn_code"])
        if transaction.transaction_type not in {"advance_received", "advance_adjusted"} and not metadata.get("uqc"):
            rows.append(["UQC_MISSING", "warning", "Unit quantity code is missing; HSN summary quantity may be incomplete.", index, transaction.reference_number, "uqc"])
        if transaction.transaction_type == "sales" and not transaction.counterparty_gstin and not transaction.place_of_supply:
            rows.append(["POS_MISSING", "error", "Place of supply is required for unregistered outward supply classification.", index, transaction.reference_number, "place_of_supply"])
        if transaction.transaction_type in {"advance_received", "advance_adjusted"} and not transaction.place_of_supply:
            rows.append(["ADVANCE_POS_MISSING", "error", "Place of supply is required for advance classification in table 11.", index, transaction.reference_number, "place_of_supply"])
        if transaction.transaction_type in {"advance_received", "advance_adjusted"} and not _transaction_has_usable_rate(transaction):
            rows.append(["ADVANCE_RATE_MISSING", "error", "Advance transaction is missing a usable GST rate.", index, transaction.reference_number, "rate"])
        if transaction.transaction_type == "advance_adjusted" and not _advance_reference_present(transaction):
            rows.append(["ADVANCE_REFERENCE_MISSING", "warning", "Advance adjustment does not reference the original advance voucher.", index, transaction.reference_number, "advance_reference"])
        if special_supply_type(transaction) in {"sez_wpay", "sez_wopay", "deemed_export"} and not transaction.counterparty_gstin:
            rows.append(["SPECIAL_SUPPLY_GSTIN_MISSING", "error", "SEZ and deemed export rows should include recipient GSTIN.", index, transaction.reference_number, "counterparty_gstin"])
        if special_supply_type(transaction) in {"export_wpay", "export_wopay"} and not _export_reference_present(transaction):
            rows.append(["EXPORT_REFERENCE_MISSING", "warning", "Export row is missing shipping bill or port-code metadata.", index, transaction.reference_number, "shipping_bill_number"])
        if ecommerce_section(transaction) and not ecommerce_gstin(transaction):
            rows.append(["ECOMMERCE_GSTIN_MISSING", "error", "E-commerce section is set but operator GSTIN is missing.", index, transaction.reference_number, "ecommerce_gstin"])
        if is_amendment_transaction(transaction) and (not original_document_number(transaction) or not original_period(transaction)):
            rows.append(["AMENDMENT_REFERENCE_MISSING", "error", "Amendment row is missing original document number or original period.", index, transaction.reference_number, "original_document_number"])
        if inferred_supply_category(transaction) in {"nil_rated", "exempt", "non_gst"} and transaction.tax_amount not in (None, Decimal("0.00")):
            rows.append(["SUPPLY_CATEGORY_CONFLICT", "warning", "Supply category indicates non-taxable treatment but tax amounts are present.", index, transaction.reference_number, "supply_category"])
    return rows


def build_gstr7_validation_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    rows = []
    duplicate_refs = {}
    for transaction in transactions:
        reference = str(transaction.reference_number or "").strip()
        if not reference:
            continue
        duplicate_refs[reference] = duplicate_refs.get(reference, 0) + 1

    for index, transaction in enumerate(transactions, start=1):
        if not str(transaction.counterparty_gstin or "").strip():
            rows.append(["DEDUCTEE_GSTIN_MISSING", "error", "Deductee GSTIN is required for GSTR-7 preparation.", index, transaction.reference_number, "counterparty_gstin"])
        if decimal_or_zero(transaction.tax_amount) <= Decimal("0.00"):
            rows.append(["ZERO_TDS_AMOUNT", "warning", "TDS amount is zero and should be reviewed before relying on this draft.", index, transaction.reference_number, "tax_amount"])
        if decimal_or_zero(transaction.total_amount) <= Decimal("0.00"):
            rows.append(["ZERO_PAYMENT_AMOUNT", "warning", "Payment amount is zero and should be reviewed for source completeness.", index, transaction.reference_number, "total_amount"])
        reference = str(transaction.reference_number or "").strip()
        if reference and duplicate_refs.get(reference, 0) > 1:
            rows.append(["DUPLICATE_TDS_DOCUMENT", "warning", "This TDS document number appears more than once in the selected period.", index, transaction.reference_number, "reference_number"])
    return rows


def taxability_code(transaction: GSTTransaction) -> int:
    category = inferred_supply_category(transaction)
    if category in {"exempt", "nil_rated"}:
        return 2
    if category == "non_gst":
        return 3
    return 1


def hsn_code(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("hsn_code") or "UNSPECIFIED")


def is_service_transaction(transaction: GSTTransaction) -> bool:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return bool(metadata.get("is_service") or metadata.get("service"))


def special_supply_type(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    value = str(metadata.get("special_supply_type") or "").strip().lower()
    if value in {"export_wpay", "export_wopay", "sez_wpay", "sez_wopay", "deemed_export"}:
        return value
    return ""


def special_supply_type_label(transaction: GSTTransaction) -> str:
    labels = {
        "export_wpay": "Export with payment",
        "export_wopay": "Export without payment",
        "sez_wpay": "SEZ with payment",
        "sez_wopay": "SEZ without payment",
        "deemed_export": "Deemed export",
    }
    return labels.get(special_supply_type(transaction), "Regular")


def ecommerce_gstin(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("ecommerce_gstin") or "").strip().upper()


def ecommerce_section(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    value = str(metadata.get("ecommerce_section") or "").strip().lower()
    if value in {"table_14", "table_15"}:
        return value
    return "table_14" if ecommerce_gstin(transaction) else ""


def is_amendment_transaction(transaction: GSTTransaction) -> bool:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return bool(
        metadata.get("is_amendment")
        or metadata.get("original_document_number")
        or metadata.get("original_document_date")
        or metadata.get("original_period")
    )


def original_document_number(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("original_document_number") or "").strip()


def original_document_date(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("original_document_date") or "").strip()


def original_period(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("original_period") or "").strip()


def original_counterparty_gstin(transaction: GSTTransaction) -> str:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return str(metadata.get("original_counterparty_gstin") or "").strip().upper()


def amendment_target_section(transaction: GSTTransaction) -> str:
    if ecommerce_gstin(transaction):
        return ecommerce_section(transaction)
    if special_supply_type(transaction):
        return special_supply_type(transaction)
    if transaction.transaction_type in {"credit_note", "debit_note"}:
        return "cdnr" if transaction.counterparty_gstin else "cdnur"
    if transaction.counterparty_gstin:
        return "b2b"
    if is_large_interstate_invoice(transaction):
        return "b2cl"
    return "b2cs"


def _transaction_has_usable_rate(transaction: GSTTransaction) -> bool:
    for component in iter_transaction_components(transaction):
        if decimal_or_zero(component.get("rate")) > Decimal("0.00"):
            return True
        taxable_value = decimal_or_zero(component.get("taxable_value"))
        tax_amount = (
            decimal_or_zero(component.get("cgst_amount"))
            + decimal_or_zero(component.get("sgst_amount"))
            + decimal_or_zero(component.get("igst_amount"))
            + decimal_or_zero(component.get("cess_amount"))
        )
        if taxable_value > Decimal("0.00") and tax_amount > Decimal("0.00"):
            return True
    return False


def _advance_reference_present(transaction: GSTTransaction) -> bool:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return bool(
        metadata.get("advance_reference")
        or metadata.get("original_advance_reference")
        or metadata.get("receipt_voucher_number")
    )


def _export_reference_present(transaction: GSTTransaction) -> bool:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    return bool(metadata.get("shipping_bill_number") or metadata.get("port_code"))


def build_rcm_contract(transaction: GSTTransaction, *, table_code: str) -> str:
    payload = {
        "invoice_tax_amounts_zero_expected": False,
        "is_reverse_charge": bool(transaction.reverse_charge),
        "liability_side": "recipient" if transaction.reverse_charge else "supplier",
        "reporting_note": "Contract metadata keeps worksheet output deterministic for downstream filing validation.",
        "table_code": table_code,
        "tax_amount_source": "invoice_amounts",
        "version": "gstr1.rcm.v2",
    }
    import json

    return json.dumps(payload, sort_keys=True)



def build_gstr3b_outward_rows(transactions: list[GSTTransaction]) -> list[list[object]]:
    sales = [transaction for transaction in transactions if transaction.transaction_type == "sales"]
    credit_notes = [transaction for transaction in transactions if transaction.transaction_type == "credit_note"]
    debit_notes = [transaction for transaction in transactions if transaction.transaction_type == "debit_note"]
    exempt_transactions = [transaction for transaction in sales if is_exempt_supply(transaction)]
    non_gst_transactions = [transaction for transaction in sales if is_non_gst_supply(transaction)]
    taxable_sales = [transaction for transaction in sales if transaction not in exempt_transactions and transaction not in non_gst_transactions]

    rows = [
        ["(a) Outward taxable supplies (other than zero rated, nil rated and exempted)", *transaction_totals_row(taxable_sales)],
        ["(b) Outward taxable supplies (zero rated)", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00"],
        ["(c) Other outward supplies (nil rated, exempted)", *transaction_totals_row(exempt_transactions)],
        ["(d) Inward supplies liable to reverse charge", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00"],
        ["(e) Non-GST outward supplies", *transaction_totals_row(non_gst_transactions)],
    ]
    if credit_notes or debit_notes:
        net_adjustments = debit_notes + credit_notes
        rows.append(["(f) Credit / debit note adjustments", *transaction_totals_row(net_adjustments)])
    return rows


def build_gstr3b_interstate_rows(transactions: list[GSTTransaction], home_state_code: str) -> list[list[object]]:
    grouped: dict[str, dict[str, Decimal]] = {}
    for transaction in transactions:
        if transaction.transaction_type != "sales":
            continue
        if transaction.counterparty_gstin:
            continue
        place_of_supply = str(transaction.place_of_supply or "").strip()
        if not place_of_supply or place_of_supply == home_state_code:
            continue
        bucket = grouped.setdefault(place_of_supply, {"taxable_value": Decimal("0.00"), "igst_amount": Decimal("0.00")})
        bucket["taxable_value"] += transaction.taxable_value or Decimal("0.00")
        bucket["igst_amount"] += transaction.igst_amount or Decimal("0.00")
    return [
        ["Unregistered", place_of_supply, format_decimal(values["taxable_value"]), format_decimal(values["igst_amount"])]
        for place_of_supply, values in sorted(grouped.items())
    ]


def build_gstr3b_exempt_rows(transactions: list[GSTTransaction], home_state_code: str) -> list[list[object]]:
    exempt = {"inter": Decimal("0.00"), "intra": Decimal("0.00")}
    non_gst = {"inter": Decimal("0.00"), "intra": Decimal("0.00")}
    for transaction in transactions:
        if transaction.transaction_type != "sales":
            continue
        bucket_key = "inter" if str(transaction.place_of_supply or "").strip() not in {"", home_state_code} else "intra"
        if is_exempt_supply(transaction):
            exempt[bucket_key] += transaction.taxable_value or Decimal("0.00")
        if is_non_gst_supply(transaction):
            non_gst[bucket_key] += transaction.taxable_value or Decimal("0.00")
    return [
        ["Nil rated / exempt outward supplies", format_decimal(exempt["inter"]), format_decimal(exempt["intra"])],
        ["Non-GST outward supplies", format_decimal(non_gst["inter"]), format_decimal(non_gst["intra"])],
    ]


def transaction_totals_row(transactions: list[GSTTransaction]) -> list[str]:
    totals = calculate_output_tax_heads(transactions)
    return [
        format_decimal(sum(((transaction.taxable_value or Decimal("0.00")) for transaction in transactions), Decimal("0.00"))),
        format_decimal(totals["igst"]),
        format_decimal(totals["cgst"]),
        format_decimal(totals["sgst"]),
        format_decimal(totals["cess"]),
        format_decimal(sum_decimal_map(totals)),
    ]


def calculate_output_tax_heads(transactions: list[GSTTransaction]) -> dict[str, Decimal]:
    totals = {"igst": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "cess": Decimal("0.00")}
    for transaction in transactions:
        sign = Decimal("-1.00") if transaction.transaction_type == "credit_note" else Decimal("1.00")
        totals["igst"] += sign * (transaction.igst_amount or Decimal("0.00"))
        totals["cgst"] += sign * (transaction.cgst_amount or Decimal("0.00"))
        totals["sgst"] += sign * (transaction.sgst_amount or Decimal("0.00"))
        totals["cess"] += sign * (transaction.cess_amount or Decimal("0.00"))
    return totals


def calculate_reconciliation_portal_head_totals(latest_run: ReconciliationRun | None) -> dict[str, dict[str, Decimal]]:
    result = {
        "eligible": {"igst": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "cess": Decimal("0.00")},
        "at_risk": {"igst": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "cess": Decimal("0.00")},
        "deferred": {"igst": Decimal("0.00"), "cgst": Decimal("0.00"), "sgst": Decimal("0.00"), "cess": Decimal("0.00")},
    }
    if latest_run is None:
        return result
    for item in latest_run.items.select_related("portal_transaction").all():
        portal = item.portal_transaction
        if portal is None:
            continue
        bucket = "at_risk"
        if item.match_status == ReconciliationItem.MatchStatus.MATCHED or item.action_status == ReconciliationItem.ActionStatus.RESOLVED:
            bucket = "eligible"
        elif item.action_status in {ReconciliationItem.ActionStatus.DEFERRED, ReconciliationItem.ActionStatus.IGNORED}:
            bucket = "deferred"
        result[bucket]["igst"] += portal.igst_amount or Decimal("0.00")
        result[bucket]["cgst"] += portal.cgst_amount or Decimal("0.00")
        result[bucket]["sgst"] += portal.sgst_amount or Decimal("0.00")
        result[bucket]["cess"] += portal.cess_amount or Decimal("0.00")
    return result


def sum_decimal_map(values: dict[str, Decimal]) -> Decimal:
    return sum(values.values(), Decimal("0.00"))


def is_exempt_supply(transaction: GSTTransaction) -> bool:
    return inferred_supply_category(transaction) in {"exempt", "nil_rated"}


def is_non_gst_supply(transaction: GSTTransaction) -> bool:
    return inferred_supply_category(transaction) == "non_gst"


def inferred_supply_category(transaction: GSTTransaction) -> str | None:
    metadata = transaction.metadata if isinstance(transaction.metadata, dict) else {}
    category = metadata.get("supply_category")
    if category in {"nil_rated", "exempt", "non_gst"}:
        return category
    if metadata.get("is_non_gst"):
        return "non_gst"
    if metadata.get("is_exempt"):
        return "exempt"
    tax_amount = decimal_or_zero(transaction.tax_amount)
    taxable_value = decimal_or_zero(transaction.taxable_value)
    if tax_amount == Decimal("0.00") and taxable_value > Decimal("0.00"):
        return "nil_rated"
    return None


def max_decimal(value: Decimal) -> Decimal:
    return value if value > Decimal("0.00") else Decimal("0.00")


def format_decimal_decimal_like(total: object, *, head: str, heads: dict[str, Decimal]) -> str:
    total_decimal = Decimal(str(total or "0.00"))
    heads_total = sum_decimal_map(heads)
    if heads_total <= Decimal("0.00"):
        return "0.00"
    ratio = heads.get(head, Decimal("0.00")) / heads_total
    return format_decimal(total_decimal * ratio)
