from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import connection
from django.test.utils import CaptureQueriesContext, override_settings
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import WorkspaceMembership, WorkspaceRole
from apps.approvals.models import ApprovalRequest
from apps.audit_logs.models import AuditLog
from apps.clients.models import Client
from apps.compliance_periods.models import CompliancePeriod
from apps.common.services.dashboard import calculate_compliance_health_score
from apps.filings.models import ReturnFiling, ReturnFilingAttempt, ReturnFilingIncidentNote
from apps.common.services.return_workbooks import (
    build_document_review_summary_rows,
    build_hsn_review_summary_rows,
    build_gstr1_b2b_rows,
    build_nil_exempt_review_rows,
)
from apps.gst_transactions.models import GSTTransaction, TransactionRemediationAssignment, TransactionRemediationFollowUp
from apps.gstins.models import GSTIN
from apps.imports.models import ImportBatch
from apps.organizations.models import Organization
from apps.reconciliation.models import ReconciliationRun
from apps.returns.models import ReturnPreparation
from apps.workspaces.models import Workspace

User = get_user_model()


def response_bytes(response):
    if getattr(response, "streaming", False):
        return b"".join(response.streaming_content)
    return response.content


@pytest.fixture
def dashboard_api_client():
    return APIClient()


@pytest.fixture
def dashboard_user(db):
    return User.objects.create_user(username="dashuser", email="dash@example.com", password="strong-pass-123")


@pytest.fixture
def dashboard_authenticated_client(dashboard_api_client, dashboard_user):
    dashboard_api_client.force_authenticate(user=dashboard_user)
    return dashboard_api_client


@pytest.fixture
def dashboard_context(dashboard_user):
    organization = Organization.objects.create(name="Dash Org", code="DASHO", created_by=dashboard_user, updated_by=dashboard_user)
    workspace = Workspace.objects.create(
        organization=organization,
        name="Dash Workspace",
        code="DASH-WS",
        created_by=dashboard_user,
        updated_by=dashboard_user,
    )
    WorkspaceMembership.objects.create(
        user=dashboard_user,
        workspace=workspace,
        role=WorkspaceRole.OWNER,
        created_by=dashboard_user,
        updated_by=dashboard_user,
    )
    client = Client.objects.create(
        workspace=workspace,
        legal_name="Dash Client Pvt Ltd",
        trade_name="Dash Client",
        client_code="DASH001",
        pan="ABCDE1234D",
        created_by=dashboard_user,
        updated_by=dashboard_user,
    )
    gstin = GSTIN.objects.create(
        client=client,
        gstin="29ABCDE1234D1Z5",
        registration_type="regular",
        state_code="29",
        created_by=dashboard_user,
        updated_by=dashboard_user,
    )
    compliance_period = CompliancePeriod.objects.create(
        gstin=gstin,
        period="2026-04",
        return_type="GSTR-3B",
        status="open",
        created_by=dashboard_user,
        updated_by=dashboard_user,
    )
    return {
        "user": dashboard_user,
        "workspace": workspace,
        "client": client,
        "gstin": gstin,
        "compliance_period": compliance_period,
    }


def create_transaction(*, context, transaction_type, reference_number):
    return GSTTransaction.objects.create(
        workspace=context["workspace"],
        client=context["client"],
        gstin=context["gstin"],
        compliance_period=context["compliance_period"],
        transaction_type=transaction_type,
        document_type="invoice",
        reference_number=reference_number,
        transaction_date="2026-04-15",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="Vendor One",
        taxable_value=Decimal("1000.00"),
        cgst_amount=Decimal("90.00"),
        sgst_amount=Decimal("90.00"),
        tax_amount=Decimal("180.00"),
        total_amount=Decimal("1180.00"),
        created_by=context["user"],
        updated_by=context["user"],
    )


@pytest.mark.django_db
def test_dashboard_summary_with_empty_data(dashboard_authenticated_client, dashboard_context):
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert data["import_summary"]["total_batches"] == 0
    assert data["reconciliation_summary"]["latest_run"] is None
    assert data["return_summary"]["gstr1"]["status"] == "not_prepared"


@pytest.mark.django_db
def test_dashboard_summary_after_imports(dashboard_authenticated_client, dashboard_context):
    ImportBatch.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        import_type="sales",
        source_type="csv",
        file_name="sales.csv",
        status="processed",
        total_rows=10,
        valid_rows=10,
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    create_transaction(context=dashboard_context, transaction_type="sales", reference_number="S-001")
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert data["import_summary"]["by_type"]["sales"] == 1
    assert data["transaction_summary"]["sales_count"] == 1


@pytest.mark.django_db
def test_dashboard_summary_after_reconciliation(dashboard_authenticated_client, dashboard_context):
    run = ReconciliationRun.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        run_type="gstr_2b_purchase",
        status="completed",
        matched_count=3,
        partial_match_count=2,
        missing_in_portal_count=1,
        duplicate_count=1,
        total_itc_at_risk=Decimal("560.00"),
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    data = response.data["data"]
    assert data["reconciliation_summary"]["latest_run"]["id"] == str(run.id)
    assert data["reconciliation_summary"]["open_issue_count"] == 4


@pytest.mark.django_db
def test_dashboard_summary_after_returns(dashboard_authenticated_client, dashboard_context):
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR1,
        status=ReturnPreparation.PreparationStatus.APPROVED,
        summary_snapshot={"outward_supplies": {"total_taxable_value": "1000.00", "total_tax_amount": "180.00"}},
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnPreparation.PreparationStatus.FILED,
        summary_snapshot={"itc_summary": {"eligible_itc": "150.00", "net_tax_payable": "30.00"}},
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        filed_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ApprovalRequest.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        entity_type="return_preparation",
        entity_id=dashboard_context["compliance_period"].id,
        status="pending",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    data = response.data["data"]
    assert data["return_summary"]["gstr1"]["status"] == "approved"
    assert data["return_summary"]["gstr3b"]["status"] == "filed"
    assert data["approval_summary"]["pending_count"] == 1


@pytest.mark.django_db
def test_dashboard_summary_query_budget(dashboard_authenticated_client, dashboard_context):
    create_transaction(context=dashboard_context, transaction_type="sales", reference_number="S-101")
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR1,
        status=ReturnPreparation.PreparationStatus.READY_FOR_REVIEW,
        summary_snapshot={"outward_supplies": {"total_taxable_value": "1000.00", "total_tax_amount": "180.00"}},
        prepared_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    with CaptureQueriesContext(connection) as queries:
        response = dashboard_authenticated_client.get(
            f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
        )

    assert response.status_code == 200
    assert len(queries) <= 25


@pytest.mark.django_db
def test_workspace_context_query_budget(dashboard_authenticated_client, dashboard_context):
    with CaptureQueriesContext(connection) as queries:
        response = dashboard_authenticated_client.get(
            f"/api/v1/workspaces/context/?workspace={dashboard_context['workspace'].id}"
        )

    assert response.status_code == 200
    assert len(queries) <= 8


@pytest.mark.django_db
@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-dashboard-cache",
            "TIMEOUT": 60,
            "KEY_PREFIX": "gst-compliance-test",
        }
    },
    CACHE_DASHBOARD_SUMMARY_SECONDS=60,
)
def test_dashboard_summary_uses_shared_cache(dashboard_authenticated_client, dashboard_context):
    cache.clear()

    with CaptureQueriesContext(connection) as first_queries:
        first_response = dashboard_authenticated_client.get(
            f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
        )

    with CaptureQueriesContext(connection) as second_queries:
        second_response = dashboard_authenticated_client.get(
            f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
        )

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert len(second_queries) < len(first_queries)
    assert len(second_queries) <= 5


@pytest.mark.django_db
@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-workspace-cache",
            "TIMEOUT": 60,
            "KEY_PREFIX": "gst-compliance-test",
        }
    },
    CACHE_WORKSPACE_CONTEXT_SECONDS=60,
)
def test_workspace_context_uses_shared_cache(dashboard_authenticated_client, dashboard_context):
    cache.clear()

    with CaptureQueriesContext(connection) as first_queries:
        first_response = dashboard_authenticated_client.get(
            f"/api/v1/workspaces/context/?workspace={dashboard_context['workspace'].id}"
        )

    with CaptureQueriesContext(connection) as second_queries:
        second_response = dashboard_authenticated_client.get(
            f"/api/v1/workspaces/context/?workspace={dashboard_context['workspace'].id}"
        )

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert len(second_queries) < len(first_queries)
    assert len(second_queries) <= 5


@pytest.mark.django_db
@override_settings(PERFORMANCE_HEADERS_ENABLED=True)
def test_dashboard_summary_includes_performance_headers(dashboard_authenticated_client, dashboard_context):
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )

    assert response.status_code == 200
    assert "X-Request-ID" in response
    assert "X-Response-Time-ms" in response
    assert "X-DB-Query-Count" in response
    assert "Server-Timing" in response


@pytest.mark.django_db
def test_error_response_includes_request_id(dashboard_api_client, dashboard_context):
    response = dashboard_api_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )

    assert response.status_code == 401
    assert response.data["status"] == "error"
    assert response.data["request_id"]
    assert response["X-Request-ID"] == response.data["request_id"]


@pytest.mark.django_db
def test_transactions_export_query_budget(dashboard_authenticated_client, dashboard_context):
    create_transaction(context=dashboard_context, transaction_type="purchase", reference_number="P-900")

    with CaptureQueriesContext(connection) as queries:
        response = dashboard_authenticated_client.get(
            f"/api/v1/exports/transactions/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
        )

    assert response.status_code == 200
    assert len(queries) <= 6


@pytest.mark.django_db
def test_dashboard_summary_includes_close_management(dashboard_authenticated_client, dashboard_context):
    assignment = TransactionRemediationAssignment.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        bucket_code="missing_hsn",
        title="Fix missing HSN rows",
        transaction_ids=[],
        filters={"status": "review"},
        status="open",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    TransactionRemediationFollowUp.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        assignment=assignment,
        follow_up_type="reminder",
        status="open",
        title="Check with reviewer",
        remind_at="2026-04-16T09:00:00Z",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert data["close_management_summary"]["assignment_count"] == 1
    assert data["close_management_summary"]["open_follow_up_count"] == 1
    assert data["close_management_summary"]["next_follow_ups"][0]["title"] == "Check with reviewer"


@pytest.mark.django_db
def test_dashboard_close_manager_view_aggregates_workspace_queues(dashboard_authenticated_client, dashboard_context):
    second_client = Client.objects.create(
        workspace=dashboard_context["workspace"],
        legal_name="Branch Client Pvt Ltd",
        trade_name="Branch Client",
        client_code="BR001",
        pan="ABCDE1234E",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    second_gstin = GSTIN.objects.create(
        client=second_client,
        gstin="29ABCDE1234E1Z5",
        registration_type="regular",
        state_code="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    second_period = CompliancePeriod.objects.create(
        gstin=second_gstin,
        period="2026-05",
        return_type="GSTR-3B",
        status="open",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    assignment = TransactionRemediationAssignment.objects.create(
        workspace=dashboard_context["workspace"],
        client=second_client,
        gstin=second_gstin,
        compliance_period=second_period,
        bucket_code="missing_uqc",
        title="Fix branch UQC rows",
        transaction_ids=[],
        filters={},
        status="open",
        assigned_to=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    TransactionRemediationFollowUp.objects.create(
        workspace=dashboard_context["workspace"],
        client=second_client,
        gstin=second_gstin,
        compliance_period=second_period,
        assignment=assignment,
        follow_up_type="reminder",
        status="open",
        title="Branch queue follow-up",
        remind_at="2026-04-16T08:00:00Z",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/close-manager/?workspace={dashboard_context['workspace'].id}"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert data["workspace"]["id"] == str(dashboard_context["workspace"].id)
    assert data["assignment_count"] >= 1
    assert len(data["queues"]) >= 1
    assert data["queues"][0]["client_name"] in {"Branch Client Pvt Ltd", "Dash Client Pvt Ltd"}


@pytest.mark.django_db
def test_dashboard_close_manager_report_returns_automation_activity(dashboard_authenticated_client, dashboard_context):
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="transaction_remediation_digest.dispatched",
        entity_type="transactionremediationdigest",
        entity_id=dashboard_context["compliance_period"].id,
        metadata={"delivery_channel": "email_preview"},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="transaction_remediation_follow_up.reminder_sent",
        entity_type="transactionremediationfollowup",
        entity_id=dashboard_context["compliance_period"].id,
        metadata={"automated": True},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="transaction_remediation_assignment.auto_escalated",
        entity_type="transactionremediationassignment",
        entity_id=dashboard_context["compliance_period"].id,
        metadata={"follow_up_id": str(dashboard_context["compliance_period"].id)},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    response = dashboard_authenticated_client.get(
        f"/api/v1/dashboard/close-manager/report/?workspace={dashboard_context['workspace'].id}&days=7"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert data["summary"]["digests_dispatched"] == 1
    assert data["summary"]["reminders_sent"] == 1
    assert data["summary"]["auto_escalations"] == 1
    assert len(data["daily"]) == 7
    assert len(data["recent_activity"]) == 3


@pytest.mark.django_db
def test_workspace_summary(dashboard_authenticated_client, dashboard_context):
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="import.uploaded",
        entity_type="ImportBatch",
        entity_id=dashboard_context["compliance_period"].id,
        metadata={},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/compliance-periods/{dashboard_context['compliance_period'].id}/workspace-summary/"
    )
    assert response.status_code == 200
    data = response.data["data"]
    assert "period_details" in data
    assert "next_recommended_action" in data
    assert isinstance(data["audit_activity"], list)


def test_health_score_calculation():
    score = calculate_compliance_health_score(
        import_summary={"by_type": {"sales": 0, "purchase": 0, "gstr_2b": 0}},
        reconciliation_summary={"open_issue_count": 5},
        return_summary={"gstr1": {"status": "not_prepared"}, "gstr3b": {"status": "not_prepared"}},
        approval_summary={"pending_count": 1},
        filing_status={"all_filed": False},
        lock_status={"is_locked": False},
    )
    assert score == 20


@pytest.mark.django_db
def test_transactions_export(dashboard_authenticated_client, dashboard_context):
    create_transaction(context=dashboard_context, transaction_type="purchase", reference_number="P-001")
    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/transactions/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&gstin={dashboard_context['gstin'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response_bytes(response)))
    sheet = workbook.active
    assert sheet["A2"].value == "purchase"
    assert sheet["C2"].value == "P-001"


@pytest.mark.django_db
def test_close_manager_report_export(dashboard_authenticated_client, dashboard_context):
    assignment = TransactionRemediationAssignment.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        bucket_code="missing_hsn",
        title="Fix missing HSN rows",
        transaction_ids=[],
        filters={"status": "review"},
        status="open",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    TransactionRemediationFollowUp.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        assignment=assignment,
        follow_up_type="reminder",
        status="open",
        title="Check with reviewer",
        remind_at="2026-04-16T09:00:00Z",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    AuditLog.objects.create(
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="transaction_remediation_follow_up.reminder_sent",
        entity_type="TransactionRemediationFollowUp",
        entity_id=assignment.id,
        after_state={"title": "Check with reviewer"},
        actor=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/close-manager-report/?workspace={dashboard_context['workspace'].id}&days=7"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    assert "Summary" in workbook.sheetnames
    assert "Client Period Queues" in workbook.sheetnames
    assert "Automation Daily" in workbook.sheetnames
    summary_sheet = workbook["Summary"]
    queues_sheet = workbook["Client Period Queues"]
    assert summary_sheet["A2"].value == "Workspace"
    assert summary_sheet["B2"].value == dashboard_context["workspace"].name
    assert queues_sheet.max_row >= 2


@pytest.mark.django_db
def test_import_errors_export(dashboard_authenticated_client, dashboard_context):
    batch = ImportBatch.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        import_type="purchase",
        source_type="csv",
        file_name="purchase.csv",
        status="failed",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    from apps.imports.models import ImportRowError

    ImportRowError.objects.create(
        import_batch=batch,
        row_number=2,
        field_name="reference_number",
        error_code="required",
        error_message="Document number is required",
        raw_row={"invoice_no": ""},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/import-errors/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&compliance_period={dashboard_context['compliance_period'].id}&import_batch={batch.id}"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    sheet = workbook.active
    assert sheet["A2"].value == "purchase.csv"
    assert sheet["D2"].value == "required"


@pytest.mark.django_db
def test_reconciliation_export(dashboard_authenticated_client, dashboard_context):
    books = create_transaction(context=dashboard_context, transaction_type="purchase", reference_number="B-001")
    portal = create_transaction(context=dashboard_context, transaction_type="gstr_2b", reference_number="G-001")
    run = ReconciliationRun.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        run_type="gstr_2b_purchase",
        status="completed",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    from apps.reconciliation.models import ReconciliationItem

    ReconciliationItem.objects.create(
        reconciliation_run=run,
        books_transaction=books,
        portal_transaction=portal,
        match_status="partial_match",
        mismatch_reason="tax_amount_mismatch",
        tax_difference=Decimal("10.00"),
        action_status="open",
        remarks="Check vendor",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/reconciliation/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&compliance_period={dashboard_context['compliance_period'].id}&run={run.id}"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    sheet = workbook.active
    assert sheet["A2"].value == str(run.id)
    assert sheet["J2"].value == "partial_match"


@pytest.mark.django_db
def test_return_summary_export(dashboard_authenticated_client, dashboard_context):
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnPreparation.PreparationStatus.APPROVED,
        summary_snapshot={
            "outward_supplies": {"outward_taxable_value": "1000.00", "outward_tax_liability": "180.00"},
            "itc_summary": {"eligible_itc": "120.00", "net_tax_payable": "60.00"},
        },
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/return-summary/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    sheet = workbook.active
    assert sheet["A2"].value == "gstr3b"
    assert sheet["E2"].value == "120.00"


@pytest.mark.django_db
def test_full_gstr1_workbook_export(dashboard_authenticated_client, dashboard_context):
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="S-001",
        transaction_date="2026-04-10",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="North Retail Pvt Ltd",
        taxable_value=Decimal("100000.00"),
        cgst_amount=Decimal("9000.00"),
        sgst_amount=Decimal("9000.00"),
        tax_amount=Decimal("18000.00"),
        total_amount=Decimal("118000.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="S-002",
        transaction_date="2026-04-12",
        counterparty_gstin="",
        counterparty_name="Cash Customer",
        taxable_value=Decimal("25000.00"),
        cgst_amount=Decimal("2250.00"),
        sgst_amount=Decimal("2250.00"),
        tax_amount=Decimal("4500.00"),
        total_amount=Decimal("29500.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="credit_note",
        document_type="credit_note",
        reference_number="CN-001",
        transaction_date="2026-04-14",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="North Retail Pvt Ltd",
        taxable_value=Decimal("1000.00"),
        cgst_amount=Decimal("90.00"),
        sgst_amount=Decimal("90.00"),
        tax_amount=Decimal("180.00"),
        total_amount=Decimal("1180.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR1,
        status=ReturnPreparation.PreparationStatus.APPROVED,
        summary_snapshot={
            "outward_supplies": {
                "b2b_taxable_value": "100000.00",
                "b2b_tax_amount": "18000.00",
                "b2c_taxable_value": "25000.00",
                "b2c_tax_amount": "4500.00",
                "credit_note_taxable_value": "1000.00",
                "credit_note_tax_amount": "180.00",
                "debit_note_taxable_value": "0.00",
                "debit_note_tax_amount": "0.00",
                "total_taxable_value": "126000.00",
                "total_tax_amount": "22680.00",
                "document_count": 3,
            }
        },
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        "/api/v1/exports/return-summary/",
        {
            "workspace": str(dashboard_context["workspace"].id),
            "client": str(dashboard_context["client"].id),
            "gstin": str(dashboard_context["gstin"].id),
            "compliance_period": str(dashboard_context["compliance_period"].id),
            "return_type": "gstr1",
            "export_mode": "full_gstr1",
        },
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    assert workbook.sheetnames == [
        "Section Summary",
        "HSN Summary",
        "Document Summary",
        "Nil Exempt",
        "Validations",
        "1_3_1_3 1 2 3 Taxpayer Details",
        "4 4 B2B",
        "5 5 B2CL (Large)",
        "6 6 Exports Deemed Exports SEZ",
        "7 7 B2CS",
        "8 8 Nil Rated Exempt Non-GST",
        "9 9 Amendments (4 5 6)",
        "10 10 CDNUR",
        "11 11 Advances and Adjustments",
        "11A Advances",
        "11B Advances",
        "12 12 HSN Summary",
        "13 13 Documents Issued",
        "14 14 Supplier ECO GSTIN-wise S",
        "14A 14A Amendments to Table 14",
        "15 15 ECO Operator GSTIN-wise B",
        "15A 15A Amendments to Table 15",
    ]
    summary_sheet = workbook["Section Summary"]
    assert summary_sheet["A2"].value == "B2B"
    assert summary_sheet["B2"].value == 1
    hsn_review_sheet = workbook["HSN Summary"]
    assert hsn_review_sheet["A2"].value == "UNSPECIFIED"
    assert hsn_review_sheet["J2"].value == 3
    document_review_sheet = workbook["Document Summary"]
    assert document_review_sheet["B2"].value == "SINV"
    assert document_review_sheet["C2"].value == 1
    assert document_review_sheet["D2"].value == 2
    nil_exempt_sheet = workbook["Nil Exempt"]
    assert nil_exempt_sheet["A1"].value == "Taxability"
    taxpayer_sheet = workbook["1_3_1_3 1 2 3 Taxpayer Details"]
    assert taxpayer_sheet["A2"].value == str(dashboard_context["gstin"].gstin)
    b2b_sheet = workbook["4 4 B2B"]
    assert b2b_sheet["B2"].value == "S-001"
    assert b2b_sheet["E2"].value == "29ABCDE9999F1Z5"
    assert b2b_sheet["J2"].value == 100000
    b2cs_sheet = workbook["7 7 B2CS"]
    assert b2cs_sheet["A2"].value == "29"
    assert b2cs_sheet["E2"].value == 25000
    note_sheet = workbook["10 10 CDNUR"]
    assert note_sheet["B2"].value == "CN-001"
    b2cl_sheet = workbook["5 5 B2CL (Large)"]
    assert b2cl_sheet["A1"].value == "Info"
    assert b2cl_sheet["A2"].value == "No rows for selected scope."
    hsn_sheet = workbook["12 12 HSN Summary"]
    assert hsn_sheet["A2"].value == "UNSPECIFIED"
    docs_sheet = workbook["13 13 Documents Issued"]
    assert docs_sheet["B2"].value == "Credit Note"
    validations_sheet = workbook["Validations"]
    assert validations_sheet["A2"].value == "HSN_MISSING"


@pytest.mark.django_db
def test_full_gstr3b_workbook_export(dashboard_authenticated_client, dashboard_context):
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="S-101",
        transaction_date="2026-04-18",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="ABC Buyer",
        taxable_value=Decimal("5000.00"),
        cgst_amount=Decimal("450.00"),
        sgst_amount=Decimal("450.00"),
        tax_amount=Decimal("900.00"),
        total_amount=Decimal("5900.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    purchase = GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="purchase",
        document_type="invoice",
        reference_number="P-001",
        transaction_date="2026-04-15",
        counterparty_gstin="29ABCDE8888F1Z5",
        counterparty_name="Vendor One",
        taxable_value=Decimal("1000.00"),
        cgst_amount=Decimal("90.00"),
        sgst_amount=Decimal("90.00"),
        tax_amount=Decimal("180.00"),
        total_amount=Decimal("1180.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    portal = GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="gstr_2b",
        document_type="invoice",
        reference_number="P-001",
        transaction_date="2026-04-15",
        counterparty_gstin="29ABCDE8888F1Z5",
        counterparty_name="Vendor One",
        taxable_value=Decimal("1000.00"),
        cgst_amount=Decimal("90.00"),
        sgst_amount=Decimal("90.00"),
        tax_amount=Decimal("180.00"),
        total_amount=Decimal("1180.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    run = ReconciliationRun.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        run_type="gstr_2b_purchase",
        status="completed",
        matched_count=1,
        total_itc_at_risk=Decimal("0.00"),
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    from apps.reconciliation.models import ReconciliationItem

    ReconciliationItem.objects.create(
        reconciliation_run=run,
        books_transaction=purchase,
        portal_transaction=portal,
        match_status="matched",
        action_status="open",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnPreparation.PreparationStatus.APPROVED,
        summary_snapshot={
            "outward_supplies": {"outward_taxable_value": "5000.00", "outward_tax_liability": "900.00"},
            "itc_summary": {
                "eligible_itc": "180.00",
                "itc_at_risk": "0.00",
                "deferred_blocked_itc": "0.00",
                "net_tax_payable": "720.00",
                "unresolved_mismatch_count": 0,
            },
            "reconciliation": {
                "latest_run_id": str(run.id),
                "matched_count": 1,
                "partial_match_count": 0,
                "missing_in_books_count": 0,
                "missing_in_portal_count": 0,
                "duplicate_count": 0,
            },
        },
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        "/api/v1/exports/return-summary/",
        {
            "workspace": str(dashboard_context["workspace"].id),
            "client": str(dashboard_context["client"].id),
            "gstin": str(dashboard_context["gstin"].id),
            "compliance_period": str(dashboard_context["compliance_period"].id),
            "return_type": "gstr3b",
            "export_mode": "full_gstr3b",
        },
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    assert workbook.sheetnames == [
        "Summary",
        "3.1 Outward Supplies",
        "3.2 Inter-State Supplies",
        "4 Eligible ITC",
        "5 Exempt Supplies",
        "5.1 Interest Late Fee",
        "6 Payment Of Tax",
        "Reconciliation Impact",
        "Source Transactions",
    ]
    summary_sheet = workbook["Summary"]
    assert summary_sheet["B2"].value == "GSTR-3B"
    outward_sheet = workbook["3.1 Outward Supplies"]
    assert outward_sheet["A2"].value == "(a) Outward taxable supplies (other than zero rated, nil rated and exempted)"
    assert outward_sheet["B2"].value == "5000.00"
    itc_sheet = workbook["4 Eligible ITC"]
    assert itc_sheet["F2"].value == "180.00"
    payment_sheet = workbook["6 Payment Of Tax"]
    assert payment_sheet["F4"].value == "720.00"
    recon_sheet = workbook["Reconciliation Impact"]
    assert recon_sheet["B2"].value == str(run.id)


@pytest.mark.django_db
def test_audit_logs_export(dashboard_authenticated_client, dashboard_context):
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        action="period.locked",
        entity_type="CompliancePeriod",
        entity_id=dashboard_context["compliance_period"].id,
        before_state={"is_locked": False},
        after_state={"is_locked": True},
        metadata={},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/audit-logs/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&compliance_period={dashboard_context['compliance_period'].id}"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    sheet = workbook.active
    assert sheet["B2"].value == dashboard_context["user"].username
    assert sheet["C2"].value == "period.locked"


@pytest.mark.django_db
def test_filing_evidence_pack_export(dashboard_authenticated_client, dashboard_context):
    prepared_return = ReturnPreparation.objects.create(
        compliance_period=dashboard_context["compliance_period"],
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnPreparation.PreparationStatus.APPROVED,
        summary_snapshot={"outward_supplies": {"outward_taxable_value": "1000.00"}},
        prepared_by=dashboard_context["user"],
        approved_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    filing = ReturnFiling.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        prepared_return=prepared_return,
        provider=ReturnFiling.Provider.WHITEBOOKS,
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnFiling.FilingStatus.SUBMITTED,
        provider_reference_id="wb-evidence-001",
        approved_by=dashboard_context["user"],
        filed_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ReturnFilingAttempt.objects.create(
        return_filing=filing,
        attempt_number=1,
        status=ReturnFilingAttempt.AttemptStatus.AWAITING_STATUS,
        request_summary={"provider_stage": "file_requested"},
        response_summary={"provider_stage": "file_requested", "next_action": "resync_for_arn_or_status"},
        triggered_by=dashboard_context["user"],
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    ReturnFilingIncidentNote.objects.create(
        return_filing=filing,
        title="Waiting for ARN",
        note="Support will resync after provider queue clears.",
        severity=ReturnFilingIncidentNote.Severity.WARNING,
        status=ReturnFilingIncidentNote.Status.OPEN,
        alert_code="delayed_arn",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    AuditLog.objects.create(
        actor=dashboard_context["user"],
        action="return_filing.submitted",
        entity_type="ReturnFiling",
        entity_id=filing.id,
        workspace_id_ref=dashboard_context["workspace"].id,
        client_id_ref=dashboard_context["client"].id,
        gstin_id_ref=dashboard_context["gstin"].id,
        compliance_period_id_ref=dashboard_context["compliance_period"].id,
        metadata={"provider_reference_id": "wb-evidence-001"},
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    response = dashboard_authenticated_client.get(
        f"/api/v1/exports/filing-evidence-pack/?workspace={dashboard_context['workspace'].id}&client={dashboard_context['client'].id}&filing={filing.id}"
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response_bytes(response)))
    assert "Filing Summary" in workbook.sheetnames
    assert "Provider Evidence" in workbook.sheetnames
    assert "Operational Alerts" in workbook.sheetnames
    assert "Incident Notes" in workbook.sheetnames
    summary_sheet = workbook["Filing Summary"]
    incident_sheet = workbook["Incident Notes"]
    assert summary_sheet["B2"].value == str(filing.id)
    assert incident_sheet["B2"].value == "Waiting for ARN"


@pytest.mark.django_db
def test_gstr1_helper_document_range_and_inferred_nil_rated(dashboard_context):
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="SI-SINV-13",
        transaction_date="2026-04-10",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="North Retail Pvt Ltd",
        taxable_value=Decimal("1000.00"),
        cgst_amount=Decimal("90.00"),
        sgst_amount=Decimal("90.00"),
        tax_amount=Decimal("180.00"),
        total_amount=Decimal("1180.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="SI-SINV-15",
        transaction_date="2026-04-11",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="North Retail Pvt Ltd",
        taxable_value=Decimal("2000.00"),
        cgst_amount=Decimal("180.00"),
        sgst_amount=Decimal("180.00"),
        tax_amount=Decimal("360.00"),
        total_amount=Decimal("2360.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )
    nil_transaction = GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="NR-001",
        transaction_date="2026-04-12",
        counterparty_gstin="",
        counterparty_name="Nil Supply Customer",
        taxable_value=Decimal("500.00"),
        cgst_amount=Decimal("0.00"),
        sgst_amount=Decimal("0.00"),
        igst_amount=Decimal("0.00"),
        tax_amount=Decimal("0.00"),
        total_amount=Decimal("500.00"),
        place_of_supply="29",
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    document_rows = build_document_review_summary_rows(
        list(GSTTransaction.objects.filter(workspace=dashboard_context["workspace"]).order_by("reference_number"))
    )
    invoice_row = next(row for row in document_rows if row[1] == "SINV")
    assert invoice_row[0] == 1
    assert invoice_row[2] == 1
    assert invoice_row[3] == 15

    nil_rows = build_nil_exempt_review_rows([nil_transaction])
    assert nil_rows == [["Nil Rated", "500.00", "0.00", "0.00", "0.00", "0.00"]]


@pytest.mark.django_db
def test_gstr1_helpers_expand_line_item_metadata(dashboard_context):
    transaction = GSTTransaction.objects.create(
        workspace=dashboard_context["workspace"],
        client=dashboard_context["client"],
        gstin=dashboard_context["gstin"],
        compliance_period=dashboard_context["compliance_period"],
        transaction_type="sales",
        document_type="invoice",
        reference_number="SI-SINV-20",
        transaction_date="2026-04-20",
        counterparty_gstin="29ABCDE9999F1Z5",
        counterparty_name="North Retail Pvt Ltd",
        taxable_value=Decimal("20500.00"),
        cgst_amount=Decimal("1812.50"),
        sgst_amount=Decimal("1812.50"),
        tax_amount=Decimal("3625.00"),
        total_amount=Decimal("24125.00"),
        place_of_supply="29",
        metadata={
            "line_items": [
                {
                    "hsn_code": "7203",
                    "is_service": True,
                    "rate": "5",
                    "taxable_value": "500.00",
                    "cgst_amount": "12.50",
                    "sgst_amount": "12.50",
                    "quantity": "0",
                    "total_amount": "525.00",
                },
                {
                    "hsn_code": "7203",
                    "is_service": False,
                    "rate": "18",
                    "taxable_value": "20000.00",
                    "cgst_amount": "1800.00",
                    "sgst_amount": "1800.00",
                    "quantity": "1110",
                    "total_amount": "23600.00",
                },
            ]
        },
        created_by=dashboard_context["user"],
        updated_by=dashboard_context["user"],
    )

    b2b_rows = build_gstr1_b2b_rows([transaction])
    assert len(b2b_rows) == 2
    assert b2b_rows[0][7] == "7203"
    assert b2b_rows[0][8] is True
    assert b2b_rows[0][9] == Decimal("500.00")
    assert b2b_rows[1][8] is False
    assert b2b_rows[1][10] == Decimal("18")

    hsn_rows = build_hsn_review_summary_rows([transaction])
    assert len(hsn_rows) == 2
    assert hsn_rows[0][0] == "7203"
    assert hsn_rows[0][1] == "N"
    assert hsn_rows[0][9] == 1
    assert hsn_rows[1][1] == "Y"
