from io import BytesIO
from pathlib import Path

import pytest
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.approvals.models import ApprovalRequest
from apps.clients.models import Client
from apps.compliance_periods.models import CompliancePeriod
from apps.gst_transactions.models import GSTTransaction
from apps.gstins.models import GSTIN
from apps.imports.models import ImportBatch, ImportRowError
from apps.reconciliation.models import ReconciliationItem, ReconciliationRun
from apps.returns.models import ReturnPreparation
from apps.workspaces.models import Workspace

SAMPLES_DIR = Path(__file__).resolve().parents[2] / "docs" / "sample-files"


def read_sample_upload(filename: str) -> SimpleUploadedFile:
    path = SAMPLES_DIR / filename
    return SimpleUploadedFile(filename, path.read_bytes(), content_type="text/csv")


@pytest.fixture
def pilot_client(db):
    return APIClient()


def login_demo(client: APIClient):
    response = client.post(
        "/api/v1/auth/token/",
        {"username": "demo_admin@example.com", "password": "demo12345"},
        format="json",
    )
    assert response.status_code == 200
    access = response.data["access"]
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {access}")
    return access


@pytest.mark.django_db
def test_phase1_pilot_workflow_end_to_end(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)

    me_response = pilot_client.get("/api/v1/auth/me/")
    assert me_response.status_code == 200
    workspace_id = me_response.data["data"]["default_workspace"]["id"]

    workspace = Workspace.objects.get(pk=workspace_id)

    client_response = pilot_client.post(
        "/api/v1/clients/",
        {
            "workspace": str(workspace.id),
            "legal_name": "Pilot Manufacturing Private Limited",
            "trade_name": "Pilot Manufacturing",
            "client_code": "PILOT001",
            "pan": "ABCDE1234K",
            "email": "pilot@example.com",
        },
        format="json",
    )
    assert client_response.status_code == 201
    client_id = client_response.data["data"]["id"]

    gstin_response = pilot_client.post(
        "/api/v1/gstins/",
        {
            "client": client_id,
            "gstin": "29ABCDE1234K1Z7",
            "registration_type": "regular",
            "state_code": "29",
        },
        format="json",
    )
    assert gstin_response.status_code == 201
    gstin_id = gstin_response.data["data"]["id"]

    period_response = pilot_client.post(
        "/api/v1/compliance-periods/",
        {
            "gstin": gstin_id,
            "period": "2026-04",
            "return_type": "GSTR-3B",
            "status": "open",
            "due_date": "2026-05-20",
        },
        format="json",
    )
    assert period_response.status_code == 201
    period_id = period_response.data["data"]["id"]

    upload_base = {
        "workspace": str(workspace.id),
        "client": client_id,
        "gstin": gstin_id,
        "compliance_period": period_id,
        "source_type": "csv",
    }
    sales_response = pilot_client.post(
        "/api/v1/imports/batches/",
        {**upload_base, "import_type": "sales", "file": read_sample_upload("sales_sample.csv")},
        format="multipart",
    )
    purchase_response = pilot_client.post(
        "/api/v1/imports/batches/",
        {**upload_base, "import_type": "purchase", "file": read_sample_upload("purchase_sample.csv")},
        format="multipart",
    )
    portal_response = pilot_client.post(
        "/api/v1/imports/batches/",
        {**upload_base, "import_type": "gstr_2b", "file": read_sample_upload("gstr_2b_sample.csv")},
        format="multipart",
    )
    assert sales_response.status_code == 201
    assert purchase_response.status_code == 201
    assert portal_response.status_code == 201

    reconciliation_response = pilot_client.post(
        "/api/v1/reconciliation/runs/",
        {
            "workspace": str(workspace.id),
            "client": client_id,
            "gstin": gstin_id,
            "compliance_period": period_id,
            "run_type": "gstr_2b_purchase",
        },
        format="json",
    )
    assert reconciliation_response.status_code == 201
    run_id = reconciliation_response.data["data"]["id"]
    run = ReconciliationRun.objects.get(pk=run_id)
    assert run.status == ReconciliationRun.RunStatus.COMPLETED

    gstr1_response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {
            "workspace": str(workspace.id),
            "client": client_id,
            "gstin": gstin_id,
            "compliance_period": period_id,
            "return_type": "gstr1",
        },
        format="json",
    )
    gstr3b_response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {
            "workspace": str(workspace.id),
            "client": client_id,
            "gstin": gstin_id,
            "compliance_period": period_id,
            "return_type": "gstr3b",
        },
        format="json",
    )
    assert gstr1_response.status_code == 200
    assert gstr3b_response.status_code == 200
    gstr3b_id = gstr3b_response.data["data"]["id"]

    request_approval_response = pilot_client.post(
        "/api/v1/approvals/",
        {
            "workspace": str(workspace.id),
            "client": client_id,
            "gstin": gstin_id,
            "compliance_period": period_id,
            "entity_type": "return_preparation",
            "entity_id": gstr3b_id,
            "comments": "Pilot review request",
        },
        format="json",
    )
    assert request_approval_response.status_code == 201
    approval_id = request_approval_response.data["data"]["id"]

    approve_approval_response = pilot_client.post(f"/api/v1/approvals/{approval_id}/approve/", {}, format="json")
    assert approve_approval_response.status_code == 200

    mark_filed_response = pilot_client.post(
        f"/api/v1/returns/{gstr3b_id}/mark-filed/",
        {"arn": "PILOTARN123"},
        format="json",
    )
    assert mark_filed_response.status_code == 200

    lock_response = pilot_client.post(f"/api/v1/compliance-periods/{period_id}/lock/", {}, format="json")
    assert lock_response.status_code == 200

    export_response = pilot_client.get(
        f"/api/v1/exports/return-summary/?workspace={workspace.id}&client={client_id}&gstin={gstin_id}&compliance_period={period_id}"
    )
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.content))
    assert workbook.active.max_row >= 2


@pytest.mark.django_db
def test_export_with_empty_data_returns_valid_xlsx(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    workspace = Workspace.objects.first()
    response = pilot_client.get(f"/api/v1/exports/transactions/?workspace={workspace.id}")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.active.max_row == 1


@pytest.mark.django_db
def test_locked_period_blocks_import_reconciliation_and_return_prep(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    workspace = Workspace.objects.first()
    client = Client.objects.filter(workspace=workspace).first()
    gstin = GSTIN.objects.filter(client=client).first()
    period = CompliancePeriod.objects.filter(gstin=gstin).first()
    period.is_locked = True
    period.save(update_fields=["is_locked"])

    import_response = pilot_client.post(
        "/api/v1/imports/batches/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "import_type": "sales",
            "source_type": "csv",
            "file": read_sample_upload("sales_sample.csv"),
        },
        format="multipart",
    )
    assert import_response.status_code == 400
    assert import_response.data["status"] == "error"

    reconciliation_response = pilot_client.post(
        "/api/v1/reconciliation/runs/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "run_type": "gstr_2b_purchase",
        },
        format="json",
    )
    assert reconciliation_response.status_code == 400

    return_response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "return_type": "gstr1",
        },
        format="json",
    )
    assert return_response.status_code == 400


@pytest.mark.django_db
def test_invalid_import_file_creates_row_errors(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    workspace = Workspace.objects.first()
    client = Client.objects.filter(workspace=workspace).first()
    gstin = GSTIN.objects.filter(client=client).first()
    period = CompliancePeriod.objects.filter(gstin=gstin).first()
    response = pilot_client.post(
        "/api/v1/imports/batches/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "import_type": "purchase",
            "source_type": "csv",
            "file": read_sample_upload("invalid_import_sample.csv"),
        },
        format="multipart",
    )
    assert response.status_code == 201
    batch = ImportBatch.objects.get(pk=response.data["data"]["id"])
    assert ImportRowError.objects.filter(import_batch=batch).exists()


@pytest.mark.django_db
def test_missing_context_and_invalid_transition_errors_are_clear(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {"return_type": "gstr1"},
        format="json",
    )
    assert response.status_code == 400
    assert response.data["status"] == "error"
    assert "workspace" in response.data["errors"]

    workspace = Workspace.objects.first()
    client = Client.objects.filter(workspace=workspace).first()
    gstin = GSTIN.objects.filter(client=client).first()
    period = CompliancePeriod.objects.filter(gstin=gstin).first()
    create_transaction = GSTTransaction.objects.create(
        workspace=workspace,
        client=client,
        gstin=gstin,
        compliance_period=period,
        transaction_type="sales",
        document_type="invoice",
        reference_number="CLEAR-001",
        transaction_date="2026-04-15",
        counterparty_gstin="29TEST1234A1Z5",
        counterparty_name="Clear Test",
        taxable_value="1000.00",
        cgst_amount="90.00",
        sgst_amount="90.00",
        tax_amount="180.00",
        total_amount="1180.00",
        created_by=workspace.created_by,
        updated_by=workspace.updated_by,
    )
    _ = create_transaction
    prepare_response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "return_type": "gstr1",
        },
        format="json",
    )
    return_id = prepare_response.data["data"]["id"]
    invalid_transition = pilot_client.post(f"/api/v1/returns/{return_id}/mark-filed/", {}, format="json")
    assert invalid_transition.status_code == 400
    assert "approved returns" in invalid_transition.data["message"].lower()


@pytest.mark.django_db
def test_reconciliation_and_return_edge_cases_without_data(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    workspace = Workspace.objects.first()
    client = Client.objects.filter(workspace=workspace).first()
    gstin = GSTIN.objects.filter(client=client).first()
    period = CompliancePeriod.objects.filter(gstin=gstin).first()

    no_purchase_response = pilot_client.post(
        "/api/v1/reconciliation/runs/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "run_type": "gstr_2b_purchase",
        },
        format="json",
    )
    assert no_purchase_response.status_code == 201
    run = ReconciliationRun.objects.get(pk=no_purchase_response.data["data"]["id"])
    assert run.status in {ReconciliationRun.RunStatus.COMPLETED, ReconciliationRun.RunStatus.FAILED}

    gstr1_response = pilot_client.post(
        "/api/v1/returns/prepare/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "return_type": "gstr1",
        },
        format="json",
    )
    assert gstr1_response.status_code == 200
    prepared = ReturnPreparation.objects.get(pk=gstr1_response.data["data"]["id"])
    assert prepared.summary_snapshot["outward_supplies"]["document_count"] >= 0


@pytest.mark.django_db
def test_duplicate_invoice_detection_and_approval_invalid_transition(pilot_client):
    call_command("seed_demo_data")
    login_demo(pilot_client)
    workspace = Workspace.objects.first()
    client = Client.objects.filter(workspace=workspace).first()
    gstin = GSTIN.objects.filter(client=client).first()
    period = CompliancePeriod.objects.filter(gstin=gstin).first()
    duplicate_upload = pilot_client.post(
        "/api/v1/imports/batches/",
        {
            "workspace": str(workspace.id),
            "client": str(client.id),
            "gstin": str(gstin.id),
            "compliance_period": str(period.id),
            "import_type": "purchase",
            "source_type": "csv",
            "file": SimpleUploadedFile(
                "duplicate.csv",
                (
                    "invoice_no,invoice_date,supplier_gstin,supplier_name,taxable_value,cgst,sgst,igst,cess,total_amount\n"
                    "DUP-001,2026-04-14,29BBBBB2222B1Z5,Vendor One,1000,90,90,0,0,1180\n"
                    "DUP-001,2026-04-15,29BBBBB2222B1Z5,Vendor One,1000,90,90,0,0,1180\n"
                ).encode("utf-8"),
                content_type="text/csv",
            ),
        },
        format="multipart",
    )
    assert duplicate_upload.status_code == 201
    batch = ImportBatch.objects.get(pk=duplicate_upload.data["data"]["id"])
    assert ImportRowError.objects.filter(import_batch=batch, error_code="duplicate_in_file").exists()

    return_preparation = ReturnPreparation.objects.create(
        compliance_period=period,
        return_type=ReturnPreparation.ReturnType.GSTR3B,
        status=ReturnPreparation.PreparationStatus.DRAFT,
        summary_snapshot={},
        created_by=workspace.created_by,
        updated_by=workspace.updated_by,
    )
    approval = ApprovalRequest.objects.create(
        workspace=workspace,
        client=client,
        gstin=gstin,
        compliance_period=period,
        entity_type="return_preparation",
        entity_id=return_preparation.id,
        status=ApprovalRequest.ApprovalStatus.PENDING,
        created_by=workspace.created_by,
        updated_by=workspace.updated_by,
    )
    invalid_approval = pilot_client.post(f"/api/v1/approvals/{approval.id}/approve/", {}, format="json")
    assert invalid_approval.status_code == 400
    assert "ready for review" in invalid_approval.data["message"].lower()
