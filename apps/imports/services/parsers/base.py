import csv
import io
import re
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from calendar import monthrange

from django.utils import timezone
from openpyxl import load_workbook

from apps.common.security import sanitize_json
from apps.imports.models import ImportRowError

GSTIN_REGEX = re.compile(r"^[0-9]{2}[A-Z0-9]{10}[A-Z0-9]{3}$")

COLUMN_ALIASES = {
    "document_number": ["invoice_no", "invoice_number", "document_number", "inv_no", "doc_no", "document_no"],
    "document_date": ["invoice_date", "document_date", "date", "inv_date", "doc_date"],
    "counterparty_gstin": ["gstin", "supplier_gstin", "recipient_gstin", "counterparty_gstin", "party_gstin", "deductee_gstin"],
    "counterparty_name": [
        "counterparty_name",
        "supplier_name",
        "recipient_name",
        "party_name",
        "customer_name",
        "vendor_name",
        "deductee_name",
    ],
    "taxable_value": ["taxable_value", "taxable_amt", "taxable_amount", "assessable_value", "deduction_base_amount"],
    "cgst_amount": ["cgst", "cgst_amount"],
    "sgst_amount": ["sgst", "sgst_amount"],
    "igst_amount": ["igst", "igst_amount"],
    "cess_amount": ["cess", "cess_amount"],
    "total_amount": ["total", "total_amount", "invoice_value", "document_value", "gross_amount", "payment_amount", "paid_amount"],
    "place_of_supply": ["place_of_supply", "pos", "state", "place_supply"],
    "reverse_charge": ["reverse_charge", "rcm", "is_reverse_charge"],
    "document_type": ["document_type", "doc_type", "invoice_type"],
    "hsn_code": ["hsn", "hsn_code", "hsn_sac", "hsn_sac_code", "sac", "sac_code"],
    "description": ["description", "item_description", "goods_description", "service_description"],
    "uqc": ["uqc", "uom", "unit", "unit_code"],
    "quantity": ["quantity", "qty", "total_qty"],
    "is_service": ["is_service", "service", "is_service_item"],
    "supply_category": ["supply_category", "taxability", "supply_type", "gst_supply_type"],
    "ecommerce_gstin": ["ecommerce_gstin", "e_commerce_gstin", "eco_gstin", "operator_gstin"],
    "rate": ["rate", "gst_rate", "tax_rate"],
    "advance_reference": ["advance_reference", "original_advance_reference", "receipt_voucher_number", "receipt_voucher_no", "receipt_voucher"],
    "special_supply_type": ["special_supply_type", "export_type", "supply_subtype", "zero_rated_type"],
    "shipping_bill_number": ["shipping_bill_number", "shipping_bill_no", "sbnum", "bill_of_export_number"],
    "shipping_bill_date": ["shipping_bill_date", "sbdt", "bill_of_export_date"],
    "port_code": ["port_code", "shipping_bill_port_code", "sbpcode"],
    "ecommerce_section": ["ecommerce_section", "eco_section", "ecom_section"],
    "original_document_number": ["original_document_number", "original_invoice_number", "original_doc_number", "oinum"],
    "original_document_date": ["original_document_date", "original_invoice_date", "original_doc_date", "oidt"],
    "original_period": ["original_period", "original_return_period", "ofp"],
    "original_counterparty_gstin": ["original_counterparty_gstin", "original_recipient_gstin", "octin"],
}


class BaseImportParser:
    transaction_type = "sales"
    default_document_type = "invoice"
    require_counterparty_gstin = False
    allowed_supply_categories = {"taxable", "nil_rated", "exempt", "non_gst"}

    def __init__(self, import_batch):
        self.import_batch = import_batch
        self._seen_document_groups = {}
        self._seen_row_signatures = set()
        self._template_mapping = {
            self._normalize_key(key): self._normalize_key(value)
            for key, value in (getattr(import_batch.import_template, "column_mapping", {}) or {}).items()
            if key and value
        }

    def read_file(self):
        if self.import_batch.source_type == "csv":
            rows = self._read_csv()
        elif self.import_batch.source_type == "excel":
            rows = self._read_excel()
        elif self.import_batch.source_type == "provider":
            rows = self._read_provider_rows()
        else:
            raise ValueError("Unsupported source type.")
        return self._apply_manual_overrides(rows)

    def validate_row(self, row_number, normalized_row, raw_row):
        issues = []
        document_group_key = self._transaction_group_key(normalized_row)
        period_exception = self._extract_period_exception(raw_row)

        if not normalized_row["document_number"]:
            issues.append(self._issue(row_number, "document_number", "required", "Document number is required.", raw_row))
        else:
            seen_group = self._seen_document_groups.get(normalized_row["document_number"])
            if seen_group and seen_group != document_group_key:
                issues.append(
                    self._issue(
                        row_number,
                        "document_number",
                        "duplicate_in_file",
                        "This document number appears more than once in the same file.",
                        raw_row,
                    )
                )
                issues.append(
                    self._issue(
                        row_number,
                        "document_number",
                        "conflicting_document_context",
                        "The same document number appears with a different date, GSTIN, or document type in the same file.",
                        raw_row,
                    )
                )
            else:
                self._seen_document_groups[normalized_row["document_number"]] = document_group_key

        if normalized_row["document_date"] is None:
            issues.append(self._issue(row_number, "document_date", "required", "Document date is required.", raw_row))
        elif not self._is_document_date_within_period(normalized_row["document_date"]) and not period_exception["allowed"]:
            issues.append(
                self._issue(
                    row_number,
                    "document_date",
                    "period_mismatch",
                    f"Document date must fall within the selected compliance period {self.import_batch.compliance_period.period}.",
                    raw_row,
                )
            )

        if self.require_counterparty_gstin and not normalized_row["counterparty_gstin"]:
            issues.append(
                self._issue(
                    row_number,
                    "counterparty_gstin",
                    "required",
                    "Counterparty GSTIN is required for this import type.",
                    raw_row,
                )
            )
        elif normalized_row["counterparty_gstin"] and not GSTIN_REGEX.match(normalized_row["counterparty_gstin"]):
            issues.append(
                self._issue(
                    row_number,
                    "counterparty_gstin",
                    "invalid_format",
                    "Counterparty GSTIN format is invalid.",
                    raw_row,
                )
            )

        if not normalized_row["counterparty_name"]:
            issues.append(
                self._issue(
                    row_number,
                    "counterparty_name",
                    "missing_name",
                    "Counterparty name is recommended for downstream review.",
                    raw_row,
                    severity=ImportRowError.Severity.WARNING,
                )
            )

        numeric_fields = ["taxable_value", "cgst_amount", "sgst_amount", "igst_amount", "cess_amount", "total_amount"]
        for field_name in numeric_fields:
            if normalized_row[field_name] is None:
                issues.append(
                    self._issue(
                        row_number,
                        field_name,
                        "invalid_number",
                        f"{field_name.replace('_', ' ').title()} must be numeric.",
                        raw_row,
                    )
                )

        if normalized_row["metadata"].get("quantity") is None and normalized_row["metadata"].get("quantity_raw"):
            issues.append(
                self._issue(
                    row_number,
                    "quantity",
                    "invalid_number",
                    "Quantity must be numeric when provided.",
                    raw_row,
                    severity=ImportRowError.Severity.WARNING,
                )
            )

        supply_category = normalized_row["metadata"].get("supply_category")
        if supply_category and supply_category not in self.allowed_supply_categories:
            issues.append(
                self._issue(
                    row_number,
                    "supply_category",
                    "invalid_choice",
                    "Supply category must be one of taxable, nil_rated, exempt, or non_gst.",
                    raw_row,
                    severity=ImportRowError.Severity.WARNING,
                )
            )

        row_signature = self._row_signature(normalized_row)
        if row_signature in self._seen_row_signatures:
            issues.append(
                self._issue(
                    row_number,
                    "document_number",
                    "duplicate_in_file",
                    "This invoice line appears to be duplicated in the same file.",
                    raw_row,
                )
            )
        else:
            self._seen_row_signatures.add(row_signature)

        return issues

    def normalize_row(self, row):
        normalized = {self._normalize_key(key): value for key, value in row.items()}
        document_number = self._coerce_string(self._pick_value(normalized, "document_number"))
        document_date = self._parse_date(self._pick_value(normalized, "document_date"))
        counterparty_gstin = self._coerce_string(self._pick_value(normalized, "counterparty_gstin")).upper()
        counterparty_name = self._coerce_string(self._pick_value(normalized, "counterparty_name"))
        taxable_value = self._parse_decimal(self._pick_value(normalized, "taxable_value"))
        cgst_amount = self._parse_decimal(self._pick_value(normalized, "cgst_amount"), default=Decimal("0.00"))
        sgst_amount = self._parse_decimal(self._pick_value(normalized, "sgst_amount"), default=Decimal("0.00"))
        igst_amount = self._parse_decimal(self._pick_value(normalized, "igst_amount"), default=Decimal("0.00"))
        cess_amount = self._parse_decimal(self._pick_value(normalized, "cess_amount"), default=Decimal("0.00"))
        total_amount = self._parse_decimal(self._pick_value(normalized, "total_amount"))
        place_of_supply = self._coerce_string(self._pick_value(normalized, "place_of_supply"))
        reverse_charge = self._parse_boolean(self._pick_value(normalized, "reverse_charge"))
        document_type = self._coerce_string(self._pick_value(normalized, "document_type")) or self.default_document_type
        hsn_code = self._coerce_string(self._pick_value(normalized, "hsn_code"))
        description = self._coerce_string(self._pick_value(normalized, "description"))
        uqc = self._coerce_string(self._pick_value(normalized, "uqc")).upper()
        quantity_raw = self._pick_value(normalized, "quantity")
        quantity = self._parse_decimal(quantity_raw)
        is_service = self._parse_boolean(self._pick_value(normalized, "is_service"))
        supply_category = self._normalize_supply_category(self._pick_value(normalized, "supply_category"))
        ecommerce_gstin = self._coerce_string(self._pick_value(normalized, "ecommerce_gstin")).upper()
        rate = self._parse_decimal(self._pick_value(normalized, "rate"), default=Decimal("0.00"))
        advance_reference = self._coerce_string(self._pick_value(normalized, "advance_reference"))
        special_supply_type = self._normalize_special_supply_type(self._pick_value(normalized, "special_supply_type"))
        shipping_bill_number = self._coerce_string(self._pick_value(normalized, "shipping_bill_number"))
        shipping_bill_date = self._parse_date(self._pick_value(normalized, "shipping_bill_date"))
        port_code = self._coerce_string(self._pick_value(normalized, "port_code")).upper()
        ecommerce_section = self._normalize_ecommerce_section(self._pick_value(normalized, "ecommerce_section"))
        original_document_number = self._coerce_string(self._pick_value(normalized, "original_document_number"))
        original_document_date = self._parse_date(self._pick_value(normalized, "original_document_date"))
        original_period = self._coerce_string(self._pick_value(normalized, "original_period"))
        original_counterparty_gstin = self._coerce_string(self._pick_value(normalized, "original_counterparty_gstin")).upper()
        tax_amount = None
        if None not in {cgst_amount, sgst_amount, igst_amount, cess_amount}:
            tax_amount = cgst_amount + sgst_amount + igst_amount + cess_amount
        if rate == Decimal("0.00") and taxable_value not in (None, Decimal("0.00")) and tax_amount not in (None, Decimal("0.00")):
            rate = (tax_amount / taxable_value) * Decimal("100.00")

        metadata = {
            "raw_columns": normalized,
            "line_items": [
                {
                    "hsn_code": hsn_code,
                    "description": description,
                    "uqc": uqc,
                    "quantity": str(quantity) if quantity is not None else None,
                    "is_service": is_service,
                    "supply_category": supply_category,
                    "ecommerce_gstin": ecommerce_gstin,
                    "rate": str(rate) if rate is not None else None,
                    "taxable_value": str(taxable_value) if taxable_value is not None else None,
                    "cgst_amount": str(cgst_amount) if cgst_amount is not None else None,
                    "sgst_amount": str(sgst_amount) if sgst_amount is not None else None,
                    "igst_amount": str(igst_amount) if igst_amount is not None else None,
                    "cess_amount": str(cess_amount) if cess_amount is not None else None,
                    "total_amount": str(total_amount) if total_amount is not None else None,
                }
            ],
        }
        if hsn_code:
            metadata["hsn_code"] = hsn_code
        if description:
            metadata["description"] = description
        if uqc:
            metadata["uqc"] = uqc
        if quantity is not None:
            metadata["quantity"] = str(quantity)
        elif quantity_raw not in (None, ""):
            metadata["quantity_raw"] = self._coerce_string(quantity_raw)
        if is_service:
            metadata["is_service"] = True
        if supply_category:
            metadata["supply_category"] = supply_category
        if ecommerce_gstin:
            metadata["ecommerce_gstin"] = ecommerce_gstin
        if rate not in (None, Decimal("0.00")):
            metadata["rate"] = str(rate)
        if advance_reference:
            metadata["advance_reference"] = advance_reference
        if special_supply_type:
            metadata["special_supply_type"] = special_supply_type
        if shipping_bill_number:
            metadata["shipping_bill_number"] = shipping_bill_number
        if shipping_bill_date is not None:
            metadata["shipping_bill_date"] = shipping_bill_date.isoformat()
        if port_code:
            metadata["port_code"] = port_code
        if ecommerce_section:
            metadata["ecommerce_section"] = ecommerce_section
        if original_document_number:
            metadata["original_document_number"] = original_document_number
        if original_document_date is not None:
            metadata["original_document_date"] = original_document_date.isoformat()
        if original_period:
            metadata["original_period"] = original_period
        if original_counterparty_gstin:
            metadata["original_counterparty_gstin"] = original_counterparty_gstin
        if original_document_number or original_document_date is not None or original_period:
            metadata["is_amendment"] = True
        period_exception = self._extract_period_exception(row)
        if period_exception["allowed"]:
            metadata["period_exception"] = {
                "allowed": True,
                "reason": period_exception["reason"],
                "category": period_exception["category"],
                "selected_period": self.import_batch.compliance_period.period,
            }

        return {
            "transaction_type": self.transaction_type,
            "document_type": document_type,
            "document_number": document_number,
            "document_date": document_date,
            "counterparty_gstin": counterparty_gstin,
            "counterparty_name": counterparty_name,
            "taxable_value": taxable_value,
            "cgst_amount": cgst_amount,
            "sgst_amount": sgst_amount,
            "igst_amount": igst_amount,
            "cess_amount": cess_amount,
            "tax_amount": tax_amount,
            "total_amount": total_amount,
            "place_of_supply": place_of_supply,
            "reverse_charge": reverse_charge,
            "metadata": metadata,
        }

    def process(self):
        rows = self.read_file()

        from apps.gst_transactions.models import GSTTransaction
        from apps.imports.models import ImportBatch

        self.import_batch.status = ImportBatch.BatchStatus.PROCESSING
        self.import_batch.processed_at = None
        self.import_batch.save(update_fields=["status", "processed_at", "updated_at"])
        self.import_batch.row_errors.all().delete()
        self.import_batch.transactions.all().delete()

        issues_to_create = []
        grouped_transactions = {}
        field_counts = Counter()
        severity_counts = Counter()
        valid_row_count = 0

        for row_number, row in rows:
            normalized_row = self.normalize_row(row)
            issues = self.validate_row(row_number, normalized_row, row)
            blocking_issues = [issue for issue in issues if issue["severity"] == ImportRowError.Severity.ERROR]

            for issue in issues:
                issues_to_create.append(
                    ImportRowError(
                        import_batch=self.import_batch,
                        row_number=row_number,
                        field_name=issue["field_name"],
                        severity=issue["severity"],
                        error_code=issue["error_code"],
                        error_message=issue["error_message"],
                        raw_row=issue["raw_row"],
                        created_by=self.import_batch.created_by,
                        updated_by=self.import_batch.updated_by,
                    )
                )
                field_counts[issue["field_name"]] += 1
                severity_counts[issue["severity"]] += 1

            if blocking_issues:
                continue

            valid_row_count += 1
            group_key = self._transaction_group_key(normalized_row)
            if group_key not in grouped_transactions:
                grouped_transactions[group_key] = self._initialize_transaction_group(normalized_row, row_number)
            else:
                self._merge_transaction_group(grouped_transactions[group_key], normalized_row, row_number)

        if issues_to_create:
            ImportRowError.objects.bulk_create(issues_to_create)
        transactions_to_create = [self._build_transaction(group_data) for group_data in grouped_transactions.values()]
        if transactions_to_create:
            GSTTransaction.objects.bulk_create(transactions_to_create)

        total_rows = len(rows)
        valid_rows = valid_row_count
        invalid_rows = len({issue.row_number for issue in issues_to_create if issue.severity == ImportRowError.Severity.ERROR})
        self.import_batch.total_rows = total_rows
        self.import_batch.valid_rows = valid_rows
        self.import_batch.invalid_rows = invalid_rows
        self.import_batch.processed_rows = total_rows
        self.import_batch.error_summary = {
            "errors": severity_counts.get(ImportRowError.Severity.ERROR, 0),
            "warnings": severity_counts.get(ImportRowError.Severity.WARNING, 0),
            "by_field": dict(field_counts),
        }
        has_manual_overrides = bool(self._manual_row_overrides()) or bool(self._discarded_row_numbers())
        self.import_batch.status = (
            ImportBatch.BatchStatus.CORRECTED if has_manual_overrides else ImportBatch.BatchStatus.PROCESSED
        )
        self.import_batch.processed_at = timezone.now()
        self.import_batch.save(
            update_fields=[
                "total_rows",
                "valid_rows",
                "invalid_rows",
                "processed_rows",
                "error_summary",
                "status",
                "processed_at",
                "updated_at",
            ]
        )

        return {
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
            "transactions_created": len(transactions_to_create),
        }

    def _read_csv(self):
        self.import_batch.file.open("rb")
        content = self.import_batch.file.read().decode("utf-8-sig")
        self.import_batch.file.close()
        reader = csv.DictReader(io.StringIO(content))
        return [(index, row) for index, row in enumerate(reader, start=2)]

    def _read_excel(self):
        self.import_batch.file.open("rb")
        workbook = load_workbook(self.import_batch.file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.import_batch.file.close()
        if not rows:
            return []
        headers = [self._coerce_string(value) for value in rows[0]]
        return [
            (index, {headers[position]: value for position, value in enumerate(row)})
            for index, row in enumerate(rows[1:], start=2)
        ]

    def _read_provider_rows(self):
        metadata = self.import_batch.source_metadata if isinstance(self.import_batch.source_metadata, dict) else {}
        rows = metadata.get("normalized_rows")
        if not isinstance(rows, list):
            return []
        return [
            (index, row)
            for index, row in enumerate((entry for entry in rows if isinstance(entry, dict)), start=2)
        ]

    def _manual_row_overrides(self):
        metadata = self.import_batch.source_metadata if isinstance(self.import_batch.source_metadata, dict) else {}
        overrides = metadata.get("manual_row_overrides")
        return overrides if isinstance(overrides, dict) else {}

    def _discarded_row_numbers(self):
        metadata = self.import_batch.source_metadata if isinstance(self.import_batch.source_metadata, dict) else {}
        discarded_rows = metadata.get("discarded_rows")
        if not isinstance(discarded_rows, list):
            return set()
        return {str(value) for value in discarded_rows}

    def _apply_manual_overrides(self, rows):
        overrides = self._manual_row_overrides()
        discarded_row_numbers = self._discarded_row_numbers()
        if not overrides:
            if not discarded_row_numbers:
                return rows
            return [(row_number, row) for row_number, row in rows if str(row_number) not in discarded_row_numbers]
        overridden_rows = []
        for row_number, row in rows:
            if str(row_number) in discarded_row_numbers:
                continue
            override = overrides.get(str(row_number))
            if isinstance(override, dict):
                overridden_rows.append((row_number, override))
            else:
                overridden_rows.append((row_number, row))
        return overridden_rows

    def _pick_value(self, normalized_row, canonical_key):
        template_key = self._template_mapping.get(canonical_key)
        if template_key and normalized_row.get(template_key) not in (None, ""):
            return normalized_row.get(template_key)
        for alias in COLUMN_ALIASES.get(canonical_key, []):
            normalized_alias = self._normalize_key(alias)
            if normalized_alias in normalized_row and normalized_row[normalized_alias] not in (None, ""):
                return normalized_row[normalized_alias]
        return normalized_row.get(canonical_key)

    def _normalize_key(self, value):
        return re.sub(r"[^a-z0-9]+", "_", self._coerce_string(value).strip().lower()).strip("_")

    def _coerce_string(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _parse_decimal(self, value, default=None):
        if value in (None, ""):
            return default
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            return None

    def _parse_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(str(value).strip(), date_format).date()
            except ValueError:
                continue
        return None

    def _parse_boolean(self, value):
        if isinstance(value, bool):
            return value
        return self._coerce_string(value).lower() in {"yes", "true", "1", "y"}

    def _normalize_supply_category(self, value):
        normalized = self._normalize_key(value)
        if normalized in {"nil", "nil_rated_supply"}:
            return "nil_rated"
        if normalized in {"nongst", "non_gst_supply"}:
            return "non_gst"
        if normalized in {"exempted"}:
            return "exempt"
        if normalized in {"taxable", "nil_rated", "exempt", "non_gst"}:
            return normalized
        return normalized or ""

    def _normalize_special_supply_type(self, value):
        normalized = self._normalize_key(value)
        aliases = {
            "export": "export_wpay",
            "export_wpay": "export_wpay",
            "wpay": "export_wpay",
            "export_with_payment": "export_wpay",
            "export_taxable": "export_wpay",
            "export_wopay": "export_wopay",
            "wopay": "export_wopay",
            "export_without_payment": "export_wopay",
            "export_exempt": "export_wopay",
            "sez_wpay": "sez_wpay",
            "sewp": "sez_wpay",
            "sez_with_payment": "sez_wpay",
            "sez_taxable": "sez_wpay",
            "sez_wopay": "sez_wopay",
            "sewop": "sez_wopay",
            "sez_without_payment": "sez_wopay",
            "deemed_export": "deemed_export",
            "deemed_exports": "deemed_export",
            "de": "deemed_export",
        }
        return aliases.get(normalized, normalized or "")

    def _normalize_ecommerce_section(self, value):
        normalized = self._normalize_key(value)
        aliases = {
            "table_14": "table_14",
            "14": "table_14",
            "supplier": "table_14",
            "table_15": "table_15",
            "15": "table_15",
            "operator": "table_15",
        }
        return aliases.get(normalized, normalized or "")

    def _issue(self, row_number, field_name, error_code, error_message, raw_row, severity=ImportRowError.Severity.ERROR):
        return {
            "row_number": row_number,
            "field_name": field_name,
            "error_code": error_code,
            "error_message": error_message,
            "severity": severity,
            "raw_row": sanitize_json(
                {str(key): "" if value is None else str(value) for key, value in raw_row.items()},
                max_items=12,
            ),
        }

    def _transaction_group_key(self, normalized_row):
        return (
            normalized_row["document_number"],
            normalized_row["document_date"].isoformat() if normalized_row["document_date"] else "",
            normalized_row["counterparty_gstin"],
            normalized_row["document_type"],
            normalized_row["transaction_type"],
        )

    def _row_signature(self, normalized_row):
        metadata = normalized_row["metadata"]
        line_item = (metadata.get("line_items") or [{}])[0]
        return (
            self._transaction_group_key(normalized_row),
            normalized_row["counterparty_name"],
            normalized_row["taxable_value"],
            normalized_row["cgst_amount"],
            normalized_row["sgst_amount"],
            normalized_row["igst_amount"],
            normalized_row["cess_amount"],
            normalized_row["total_amount"],
            normalized_row["place_of_supply"],
            normalized_row["reverse_charge"],
            line_item.get("hsn_code"),
            line_item.get("description"),
            line_item.get("uqc"),
            line_item.get("quantity"),
            line_item.get("is_service"),
            line_item.get("supply_category"),
            line_item.get("ecommerce_gstin"),
            line_item.get("rate"),
            metadata.get("advance_reference"),
        )

    def _initialize_transaction_group(self, normalized_row, row_number):
        return {
            "normalized_row": normalized_row,
            "source_row_numbers": [row_number],
            "metadata_rows": [row_number],
        }

    def _merge_transaction_group(self, group_data, normalized_row, row_number):
        base_row = group_data["normalized_row"]
        base_row["taxable_value"] = self._add_decimals(base_row["taxable_value"], normalized_row["taxable_value"])
        base_row["cgst_amount"] = self._add_decimals(base_row["cgst_amount"], normalized_row["cgst_amount"])
        base_row["sgst_amount"] = self._add_decimals(base_row["sgst_amount"], normalized_row["sgst_amount"])
        base_row["igst_amount"] = self._add_decimals(base_row["igst_amount"], normalized_row["igst_amount"])
        base_row["cess_amount"] = self._add_decimals(base_row["cess_amount"], normalized_row["cess_amount"])
        base_row["tax_amount"] = self._add_decimals(base_row["tax_amount"], normalized_row["tax_amount"])
        base_row["total_amount"] = self._add_decimals(base_row["total_amount"], normalized_row["total_amount"])
        if not base_row["counterparty_name"] and normalized_row["counterparty_name"]:
            base_row["counterparty_name"] = normalized_row["counterparty_name"]
        if not base_row["place_of_supply"] and normalized_row["place_of_supply"]:
            base_row["place_of_supply"] = normalized_row["place_of_supply"]
        base_row["reverse_charge"] = base_row["reverse_charge"] or normalized_row["reverse_charge"]

        base_metadata = base_row["metadata"]
        base_metadata.setdefault("line_items", [])
        base_metadata["line_items"].extend(normalized_row["metadata"].get("line_items", []))
        base_metadata.setdefault("source_rows", []).append(row_number)
        base_metadata["aggregated_line_count"] = len(base_metadata["line_items"])

        for key in ("hsn_code", "description", "uqc", "quantity", "is_service", "supply_category", "ecommerce_gstin", "rate", "advance_reference"):
            existing_value = base_metadata.get(key)
            incoming_value = normalized_row["metadata"].get(key)
            if not incoming_value:
                continue
            if not existing_value:
                base_metadata[key] = incoming_value
            elif existing_value != incoming_value:
                base_metadata.pop(key, None)
                base_metadata.setdefault("mixed_fields", [])
                if key not in base_metadata["mixed_fields"]:
                    base_metadata["mixed_fields"].append(key)

        group_data["source_row_numbers"].append(row_number)
        group_data["metadata_rows"].append(row_number)

    def _build_transaction(self, group_data):
        from apps.gst_transactions.models import GSTTransaction

        normalized_row = group_data["normalized_row"]
        metadata = normalized_row["metadata"]
        metadata["source_rows"] = group_data["source_row_numbers"]
        metadata["aggregated_line_count"] = len(metadata.get("line_items", []))

        return GSTTransaction(
            workspace=self.import_batch.workspace,
            client=self.import_batch.client,
            gstin=self.import_batch.gstin,
            compliance_period=self.import_batch.compliance_period,
            transaction_type=normalized_row["transaction_type"],
            document_type=normalized_row["document_type"],
            reference_number=normalized_row["document_number"],
            transaction_date=normalized_row["document_date"],
            counterparty_gstin=normalized_row["counterparty_gstin"],
            counterparty_name=normalized_row["counterparty_name"],
            taxable_value=normalized_row["taxable_value"],
            cgst_amount=normalized_row["cgst_amount"] or Decimal("0.00"),
            sgst_amount=normalized_row["sgst_amount"] or Decimal("0.00"),
            igst_amount=normalized_row["igst_amount"] or Decimal("0.00"),
            cess_amount=normalized_row["cess_amount"] or Decimal("0.00"),
            tax_amount=normalized_row["tax_amount"] or Decimal("0.00"),
            total_amount=normalized_row["total_amount"],
            place_of_supply=normalized_row["place_of_supply"],
            reverse_charge=normalized_row["reverse_charge"],
            import_batch=self.import_batch,
            metadata=metadata,
            created_by=self.import_batch.created_by,
            updated_by=self.import_batch.updated_by,
        )

    def _add_decimals(self, left, right):
        return (left or Decimal("0.00")) + (right or Decimal("0.00"))

    def _is_document_date_within_period(self, document_date):
        if document_date is None:
            return False
        period_value = getattr(self.import_batch.compliance_period, "period", "")
        try:
            year, month = [int(part) for part in str(period_value).split("-", 1)]
        except (TypeError, ValueError):
            return True

        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
        return start_date <= document_date <= end_date

    def _extract_period_exception(self, raw_row):
        if not isinstance(raw_row, dict):
            return {"allowed": False, "reason": "", "category": ""}

        allow_value = self._coerce_string(raw_row.get("__allow_period_exception")).lower()
        allowed = allow_value in {"true", "1", "yes", "y"}
        reason = self._coerce_string(raw_row.get("__period_exception_reason"))
        category = self._coerce_string(raw_row.get("__period_exception_category"))

        if not allowed or not reason:
          return {"allowed": False, "reason": "", "category": ""}
        return {
            "allowed": True,
            "reason": reason,
            "category": category,
        }
